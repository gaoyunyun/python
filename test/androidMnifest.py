import copy
import os
import re

from androguard.core.bytecodes.axml import AXMLPrinter
from androguard.misc import AnalyzeAPK
import requests
from bs4 import BeautifulSoup
import json
import lookup
import xlwt
import pandas as pd

# 唯一标识 ：记录总个数用
from openpyxl import load_workbook

uniqueId = 1;
软件设计17年上 = {
    "code": 0,
    "list": [{
        "tihao": 1,
        "q": "CPU执行算术运算或者逻辑运算时，常将源操作数和结果暂存在（ ）中。",
        "a1": "A. 程序计数器 (PC)",
        "a2": "B. 累加器 (AC)",
        "a3": "C. 指令寄存器 (IR)",
        "a4": "D. 地址寄存器 (AR)",
        "sl": "B"
    }, {
        "tihao": 3,
        "q": "计算机系统中常用的输入/输出控制方式有无条件传送、中断、程序查询和 DMA方式等。当采用（）方式时，不需要CPU 执行程序指令来传送数据。",
        "a1": "A. 中断",
        "a2": "B. 程序查询",
        "a3": "C. 无条件传送",
        "a4": "D. DMA",
        "sl": "D"
    }, {
        "tihao": 4,
        "q": "某系统由下图所示的冗余部件构成。若每个部件的千小时可靠度都为R ，则该系统的千小时可靠度为（）。",
        "a1": "A. (1-R3)(1-R2)",
        "a2": "B. (1-(1-R)3)(1-(1-R)2)",
        "a3": "C. (1-R3)+(1-R2)",
        "a4": "D. (1-(1-R)3)+(1-(1-R)2)",
        "sl": "B"
    }, {
        "tihao": 5,
        "q": "已知数据信息为16位，最少应附加（）位校验位，才能实现海明码纠错。",
        "a1": "A. 3",
        "a2": "B. 4",
        "a3": "C. 5",
        "a4": "D. 6",
        "sl": "C"
    }, {
        "tihao": 6,
        "q": "以下关于Cache (高速缓冲存储器)的叙述中，不正确的是（）",
        "a1": "A. Cache的设置扩大了主存的容量",
        "a2": "B. Cache的内容是主存部分内容的拷贝",
        "a3": "C. Cache的命中率并不随其容量增大线性地提高",
        "a4": "D. Cache位于主存与CPU之间",
        "sl": "A"
    }, {
        "tihao": 7,
        "q": "HTTPS使用（）协议对报文进行封装。",
        "a1": "A. SSH",
        "a2": "B. SSL",
        "a3": "C. SHA-1",
        "a4": "D. SET",
        "sl": "B"
    }, {
        "tihao": 8,
        "q": "以下加密算法中适合对大量的明文消息进行加密传输的是（ ）。",
        "a1": "A. RSA",
        "a2": "B. SHA-1",
        "a3": "C. MD5",
        "a4": "D. RC5",
        "sl": "D"
    }, {
        "tihao": 9,
        "q": "假定用户A、B 分别在I1和I2两个CA处取得了各自的证书，下面（）是A、B 互信的必要条件。",
        "a1": "A. A、B互换私钥",
        "a2": "B. A、B互换公钥",
        "a3": "C. I1、I2互换私钥",
        "a4": "D. I1、I2互换公钥",
        "sl": "D"
    }, {
        "tihao": 10,
        "q": "甲软件公司受乙企业委托安排公司软件设计师开发了信息系统管理软件，由于在委托开发合同中未对软件著作权归属作出明确的约定，所以该信息系统管理软件的著作权由( )享有。",
        "a1": "A. 甲",
        "a2": "B. 乙",
        "a3": "C. 甲与乙共同",
        "a4": "D. 软件设计师",
        "sl": "A"
    }, {
        "tihao": 11,
        "q": "根据我国商标法，下列商品中必须使用注册商标的是（）。",
        "a1": "A. 医疗仪器",
        "a2": "B. 墙壁涂料",
        "a3": "C. 无糖食品",
        "a4": "D. 烟草制品",
        "sl": "D"
    }, {
        "tihao": 12,
        "q": "甲、乙两人在同一天就同样的发明创造提交了专利申请，专利局将分别向各申请人通报有关情况，并提出多种可能采用的解决办法。下列说法中，不可能采用( )。",
        "a1": "A. 甲、乙作为共同申请人",
        "a2": "B. 甲或乙一方放弃权利并从另一方得到适当的补偿",
        "a3": "C. 甲、乙都不授予专利权",
        "a4": "D. 甲、乙都授予专利权",
        "sl": "D"
    }, {
        "tihao": 13,
        "q": "数字语音的采样频率定义为8kHz，这是因为( )。",
        "a1": "A. 语音信号定义的频率最高值为4kHz",
        "a2": "B. 语音信号定义的频率最高值为8kHz",
        "a3": "C. 数字语音转输线路的带宽只有8kHz",
        "a4": "D. 一般声卡的采样频率最高为每秒8k次",
        "sl": "A"
    }, {
        "tihao": 14,
        "q": "使用图像扫描仪以300DPI的分辨率扫描一幅3×4英寸的图片，可以得到( )像素的数字图像。",
        "a1": "A. 300×300",
        "a2": "B. 300×400",
        "a3": "C. 900×4",
        "a4": "D. 900×1200",
        "sl": "D"
    }, {
        "tihao": 15,
        "q": "在采用结构化开发方法进行软件开发时，设计阶段接口设计主要依据需求分析阶段的(15)。接口设计的任务主要是(16)。",
        "a1": "A. 数据流图",
        "a2": "B. E-R图",
        "a3": "C. 状态-迁移图",
        "a4": "D. 加工规格说明",
        "sl": "A"
    }, {
        "tihao": 16,
        "q": "在采用结构化开发方法进行软件开发时，设计阶段接口设计主要依据需求分析阶段的(15)。接口设计的任务主要是(16)。",
        "a1": "A. 定义软件的主要结构元素及其之间的关系",
        "a2": "B. 确定软件涉及的文件系统的结构及数据库的表结构",
        "a3": "C. 描述软件与外部环境之间的交互关系，软件内模块之间的调用关系",
        "a4": "D. 确定软件各个模块内部的算法和数据结构",
        "sl": "C"
    }, {
        "tihao": 17,
        "q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示活动的持续时间(天)，则完成该项目的最早时间为(17)天。活动BD和HK最早可以从第(18)天开始。(活动AB、AE和AC最早从第1天开始)",
        "a1": "A. 17",
        "a2": "B. 18",
        "a3": "C. 19",
        "a4": "D. 20",
        "sl": "D"
    }, {
        "tihao": 18,
        "q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示活动的持续时间(天)，则完成该项目的最早时间为(17)天。活动BD和HK最早可以从第(18)天开始。(活动AB、AE和AC最早从第1天开始)",
        "a1": "A. 3和10",
        "a2": "B. 4和11",
        "a3": "C. 3和9",
        "a4": "D. 4和10",
        "sl": "B"
    }, {
        "tihao": 19,
        "q": "在进行软件开发时，采用无主程序员的开发小组，成员之间相互平等;而主程序员负责制的开发小组，由一个主程序员和若干成员组成，成员之间没有沟通。在一个由8名开发人员构成的小组中，无主程序员组和主程序员组的沟通路径分别是()。",
        "a1": "A. 32和8",
        "a2": "B. 32和7",
        "a3": "C. 28和8",
        "a4": "D. 28和7",
        "sl": "D"
    }, {
        "tihao": 20,
        "q": "在高级语言源程序中，常需要用户定义的标识符为程序中的对象命名，常见的命名对象有( )①关键字（或保留字）②变量③函数④数据类型⑤注释",
        "a1": "A. ①②③",
        "a2": "B. ②③④",
        "a3": "C. ①③⑤",
        "a4": "D. ②④⑤",
        "sl": "B"
    }, {
        "tihao": 21,
        "q": "在仅由字符a、b构成的所有字符串中，其中以b结尾的字符串集合可用正则表达式为( )。",
        "a1": "A. (b|ab)*b",
        "a2": "B. (ab*)*b",
        "a3": "C. a*b*b",
        "a4": "D. (a|b)*b",
        "sl": "D"
    }, {
        "tihao": 22,
        "q": "在以阶段划分的编译过程中，判断程序语句的形式是否正确属于() 阶段的工作。",
        "a1": "A. 词法分析",
        "a2": "B. 语法分析",
        "a3": "C. 语义分析",
        "a4": "D. 代码生成",
        "sl": "B"
    }, {
        "tihao": 23,
        "q": "某文件管理系统在磁盘上建立了位示图(bitmap) ，记录磁盘的使用情况。若计算机系统的字长为32 位，磁盘的容量为300GB ，物理块的大小为4MB ，那么位示图的大小需要( )个字。",
        "a1": "A. 1200",
        "a2": "B. 2400",
        "a3": "C. 6400",
        "a4": "D. 9600",
        "sl": "B"
    }, {
        "tihao": 24,
        "q": "某系统中有3个并发进程竞争资源R，每个进程都需要5个R，那么至少有()个R，才能保证系统不会发生死锁。",
        "a1": "A. 12",
        "a2": "B. 13",
        "a3": "C. 14",
        "a4": "D. 15",
        "sl": "B"
    }, {
        "tihao": 25,
        "q": "某计算机系统页面大小为4K ，进程的页面变换表如下所示。若进程的逻辑地址为2D16H。该地址经过变换后，其物理地址应为( )",
        "a1": "A. 2048H",
        "a2": "B. 4096H",
        "a3": "C. 4D16H",
        "a4": "D. 6D16H",
        "sl": "C"
    }, {
        "tihao": 26,
        "q": "进程P1、P2 、P3、P4 和P5的前趋图如下所示:若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，需要设置5个信号量S1、S2、S3、S4和S5，且信号量S1~S5的初值都等于零。如下的进程执行图中a和b处应分别填写(26)；c和d处应分别填写(27)；e和f处应分别填写(28) 。",
        "a1": "A. V(S1)和P(S2)V(S3)",
        "a2": "B. P(S1)和V(S2)V(S3)",
        "a3": "C. V(S1)和V(S2)V(S3)",
        "a4": "D. P(S1)和P(S2)V(S3)",
        "sl": "B"
    }, {
        "tihao": 27,
        "q": "进程P1、P2 、P3、P4 和P5的前趋图如下所示:若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，需要设置5个信号量S1、S2、S3、S4和S5，且信号量S1~S5的初值都等于零。如下的进程执行图中a和b处应分别填写(26)；c和d处应分别填写(27)；e和f处应分别填写(28) 。",
        "a1": "A. P(S2)和P(S4)",
        "a2": "B. V(S2)和P(S4)",
        "a3": "C. P(S2)和V(S3)",
        "a4": "D. V(S2)和V(S4)",
        "sl": "C"
    }, {
        "tihao": 28,
        "q": "进程P1、P2 、P3、P4 和P5的前趋图如下所示:若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，需要设置5个信号量S1、S2、S3、S4和S5，且信号量S1~S5的初值都等于零。如下的进程执行图中a和b处应分别填写(26)；c和d处应分别填写(27)；e和f处应分别填写(28) 。",
        "a1": "A. P(S4)和V(S5)",
        "a2": "B. V(S5)和P(S4)",
        "a3": "C. V(S4)和P(S5)",
        "a4": "D. V(S4)和V(S5)",
        "sl": "A"
    }, {
        "tihao": 29,
        "q": "以下关于螺旋模型的叙述中，不正确的是( )。",
        "a1": "A. 它是风险驱动的，要求开发人员必须具有丰富的风险评估知识和经验",
        "a2": "B. 它可以降低过多测试或测试不足带来的风险",
        "a3": "C. 它包含维护周期，因此维护和开发之间没有本质区别",
        "a4": "D. 它不适用于大型软件开发",
        "sl": "D"
    }, {
        "tihao": 30,
        "q": "以下关于极限编程(XP) 中结对编程的叙述中，不正确的是( )。",
        "a1": "A. 支持共同代码拥有和共同对系统负责",
        "a2": "B. 承担了非正式的代码审查过程",
        "a3": "C. 代码质量更高",
        "a4": "D. 编码速度更快",
        "sl": "D"
    }, {
        "tihao": 31,
        "q": "以下关于C/S (客户机/服务器)体系结构的优点的叙述中，不正确的是（）。",
        "a1": "A. 允许合理地划分三层的功能，使之在逻辑上保持相对独立性",
        "a2": "B. 允许各层灵活地选用平台和软件",
        "a3": "C. 各层可以选择不同的开发语言进行并行开发",
        "a4": "D. 系统安装、修改和维护均只在服务器端进行",
        "sl": "D"
    }, {
        "tihao": 32,
        "q": "在设计软件的模块结构时， ( )不能改进设计质量。",
        "a1": "A. 尽量减少高扇出结构",
        "a2": "B. 模块的大小适中",
        "a3": "C. 将具有相似功能的模块合并",
        "a4": "D. 完善模块的功能",
        "sl": "C"
    }, {
        "tihao": 33,
        "q": "模块A、B和C有相同的程序块，块内的语句之间没有任何联系，现把该程序块取出来，形成新的模块D，则模块D的内聚类型为(33)内聚。以下关于该内聚类型的叙述中，不正确的是(34)。",
        "a1": "A. 巧合",
        "a2": "B. 逻辑",
        "a3": "C. 时间",
        "a4": "D. 过程",
        "sl": "A"
    }, {
        "tihao": 34,
        "q": "模块A、B和C有相同的程序块，块内的语句之间没有任何联系，现把该程序块取出来，形成新的模块D，则模块D的内聚类型为(33)内聚。以下关于该内聚类型的叙述中，不正确的是(34)。",
        "a1": "A. 具有最低的内聚性",
        "a2": "B. 不易修改和维护",
        "a3": "C. 不易理解",
        "a4": "D. 不影响模块间的耦合关系",
        "sl": "D"
    }, {
        "tihao": 35,
        "q": "对下图所示的程序流程图进行语句覆盖测试和路径覆盖测试，至少需要(35)个测试用例。采用McCabe度量法计算其环路复杂度为(36)。",
        "a1": "A. 2和3",
        "a2": "B. 2和4",
        "a3": "C. 2和5",
        "a4": "D. 2和6",
        "sl": "B"
    }, {
        "tihao": 36,
        "q": "对下图所示的程序流程图进行语句覆盖测试和路径覆盖测试，至少需要(35)个测试用例。采用McCabe度量法计算其环路复杂度为(36)。",
        "a1": "A. 1",
        "a2": "B. 2",
        "a3": "C. 3",
        "a4": "D. 4",
        "sl": "D"
    }, {
        "tihao": 37,
        "q": "在面向对象方法中，两个及以上的类作为一个类的父类时，称为(37)，使用它可能造成子类中存在(38)的成员。",
        "a1": "A. 多重继承",
        "a2": "B. 多态",
        "a3": "C. 封装",
        "a4": "D. 层次继承",
        "sl": "A"
    }, {
        "tihao": 38,
        "q": "在面向对象方法中，两个及以上的类作为一个类的父类时，称为(37)，使用它可能造成子类中存在(38)的成员。",
        "a1": "A. 动态",
        "a2": "B. 私有",
        "a3": "C. 公共",
        "a4": "D. 二义性",
        "sl": "D"
    }, {
        "tihao": 39,
        "q": "采用面向对象方法进行软件开发，在分析阶段，架构师主要关注系统的( )。",
        "a1": "A. 技术",
        "a2": "B. 部署",
        "a3": "C. 实现",
        "a4": "D. 行为",
        "sl": "D"
    }, {
        "tihao": 40,
        "q": "在面向对象方法中,多态指的是( )。",
        "a1": "A. 客户类无需知道所调用方法的特定子类的实现",
        "a2": "B. 对象动态地修改类",
        "a3": "C. 一个对象对应多张数据库表",
        "a4": "D. 子类只能够覆盖父类中非抽象的方法",
        "sl": "A"
    }, {
        "tihao": 41,
        "q": "",
        "a1": "A. 序列图",
        "a2": "B. 状态图",
        "a3": "C. 通信图",
        "a4": "D. 活动图",
        "sl": "C"
    }, {
        "tihao": 42,
        "q": "",
        "a1": "A. 类",
        "a2": "B. 对象",
        "a3": "C. 流名称",
        "a4": "D. 消息",
        "sl": "B"
    }, {
        "tihao": 43,
        "q": "",
        "a1": "A. 类",
        "a2": "B. 对象",
        "a3": "C. 流名称",
        "a4": "D. 消息",
        "sl": "D"
    }, {
        "tihao": 44,
        "q": "下图所示为观察者(Observer)模式的抽象示意图，其中( 44)知道其观察者，可以有任何多个观察者观察同一个目标；提供注册和删除观察者对象的接口。此模式体现的最主要的特征是(45 )。",
        "a1": "A. Subject",
        "a2": "B. Observer",
        "a3": "C. ConcreteSubject",
        "a4": "D. ConcreteObserver",
        "sl": "A"
    }, {
        "tihao": 45,
        "q": "下图所示为观察者(Observer)模式的抽象示意图，其中( 44)知道其观察者，可以有任何多个观察者观察同一个目标；提供注册和删除观察者对象的接口。此模式体现的最主要的特征是(45 )。",
        "a1": "A. 类应该对扩展开放，对修改关闭",
        "a2": "B. 使所要交互的对象尽量松耦合",
        "a3": "C. 组合优先于继承使用",
        "a4": "D. 仅与直接关联类交互",
        "sl": "B"
    }, {
        "tihao": 46,
        "q": "装饰器 (Decorator) 模式用于 (46);外观 (Facade) 模式用于(47)。①将一个对象加以包装以给客户提供其希望的另外一个接口②将一个对象加以包装以提供一些额外的行为③将一个对象加以包装以控制对这个对象的访问④将一系列对象加以包装以简化其接口",
        "a1": "A. ①",
        "a2": "B. ②",
        "a3": "C. ③",
        "a4": "D. ④",
        "sl": "B"
    }, {
        "tihao": 47,
        "q": "装饰器 (Decorator) 模式用于 (46);外观 (Facade) 模式用于(47)。①将一个对象加以包装以给客户提供其希望的另外一个接口②将一个对象加以包装以提供一些额外的行为③将一个对象加以包装以控制对这个对象的访问④将一系列对象加以包装以简化其接口",
        "a1": "A. ①",
        "a2": "B. ②",
        "a3": "C. ③",
        "a4": "D. ④",
        "sl": "D"
    }, {
        "tihao": 48,
        "q": "某确定的有限自动机 (DFA) 的状态转换图如下图所示 (A 是初态，D、E 是终态)，则该 DFA 能识别 ( )",
        "a1": "A. 00110",
        "a2": "B. 10101",
        "a3": "C. 11100",
        "a4": "D. 11001",
        "sl": "C"
    }, {
        "tihao": 49,
        "q": "函数main()、f()的定义如下所示，调用函数们f()时，第一个参数采用传值(call by value) 方式，第二个参数采用传引用(call by reference) 方式，main() 函数中 'print(x)' 执行后输出的值为( )。",
        "a1": "A. 11",
        "a2": "B. 40",
        "a3": "C. 45",
        "a4": "D. 70",
        "sl": "B"
    }, {
        "tihao": 50,
        "q": "下图为一个表达式的语法树，该表达式的后缀形式为 ( )",
        "a1": "A. x5y+*a/b-",
        "a2": "B. x5yab*+/-",
        "a3": "C. -/*x+5yab",
        "a4": "D. x5*y+a/b-",
        "sl": "A"
    }, {
        "tihao": 51,
        "q": "若事务T1对数据D1加了共享锁，事务T2、T3分别对数据D2、D3加了排它锁，则事务T1对数据(51) ；事务T2对数据(52)。",
        "a1": "A. D2、D3加排它锁都成功",
        "a2": "B. D2、D3加共享锁都成功",
        "a3": "C. D2加共享锁成功，D3加排它锁失败",
        "a4": "D. D2、D3加排它锁和共享锁都失败",
        "sl": "D"
    }, {
        "tihao": 52,
        "q": "若事务T1对数据D1加了共享锁，事务T2、T3分别对数据D2、D3加了排它锁，则事务T1对数据(51) ；事务T2对数据(52)。",
        "a1": "A. D1、D3加共享锁都失败",
        "a2": "B. D1、D3加共享锁都成功",
        "a3": "C. D1加共享锁成功 ，D3加排它锁失败",
        "a4": "D. D1加排它锁成功 ，D3加共享锁失败",
        "sl": "C"
    }, {
        "tihao": 53,
        "q": "假设关系R<U,F>,U= {A1,A2, A3}，F = {A1A3 →A2,A1A2 →A3}，则关系R的各候选关键字中必定含有属性(  )。",
        "a1": "A. A1",
        "a2": "B. A2",
        "a3": "C. A3",
        "a4": "D. A2 A3",
        "sl": "A"
    }, {
        "tihao": 54,
        "q": "在某企业的工程项目管理系统的数据库中供应商关系Supp、项目关系Proj和零件关系Part的E-R模型和关系模式如下:Supp（供应商号,供应商名,地址,电话）Proj（项目号,项目名,负责人,电话）Part（零件号,零件名）其中，每个供应商可以为多个项目供应多种零件，每个项目可由多个供应商供应多种零件。SP_P需要生成一个独立的关系模式，其联系类型为（54）给定关系模式SP_P（供应商号,项目号,零件号,数量）查询至少供应了3个项目（包含3项）的供应商，输出其供应商号和供应零件数量的总和，并按供应商号降序排列。SELECT 供应商号，SUM（数量） FROM (55) GROUP BY 供应商号(56) ORDER BY 供应商号DESC;",
        "a1": "A. *:*:*",
        "a2": "B. 1:*:*",
        "a3": "C. 1:1:*",
        "a4": "D. 1:1:1",
        "sl": "A"
    }, {
        "tihao": 55,
        "q": "在某企业的工程项目管理系统的数据库中供应商关系Supp、项目关系Proj和零件关系Part的E-R模型和关系模式如下:Supp（供应商号,供应商名,地址,电话）Proj（项目号,项目名,负责人,电话）Part（零件号,零件名）其中，每个供应商可以为多个项目供应多种零件，每个项目可由多个供应商供应多种零件。SP_P需要生成一个独立的关系模式，其联系类型为（54）给定关系模式SP_P（供应商号,项目号,零件号,数量）查询至少供应了3个项目（包含3项）的供应商，输出其供应商号和供应零件数量的总和，并按供应商号降序排列。SELECT 供应商号，SUM（数量） FROM (55) GROUP BY 供应商号(56) ORDER BY 供应商号DESC;",
        "a1": "A. Supp",
        "a2": "B. Proj",
        "a3": "C. Part",
        "a4": "D. SP_P",
        "sl": "D"
    }, {
        "tihao": 56,
        "q": "在某企业的工程项目管理系统的数据库中供应商关系Supp、项目关系Proj和零件关系Part的E-R模型和关系模式如下:Supp（供应商号,供应商名,地址,电话）Proj（项目号,项目名,负责人,电话）Part（零件号,零件名）其中，每个供应商可以为多个项目供应多种零件，每个项目可由多个供应商供应多种零件。SP_P需要生成一个独立的关系模式，其联系类型为（54）给定关系模式SP_P（供应商号,项目号,零件号,数量）查询至少供应了3个项目（包含3项）的供应商，输出其供应商号和供应零件数量的总和，并按供应商号降序排列。SELECT 供应商号，SUM（数量） FROM (55) GROUP BY 供应商号(56) ORDER BY 供应商号DESC;",
        "a1": "A. HAVING COUNT(项目号)>2",
        "a2": "B. WHERE COUNT(项目号)>2",
        "a3": "C. HAVING COUNT(DISTINCT(项目号))>2",
        "a4": "D. WHERE COUNT(DISTINCT(项目号))>3",
        "sl": "C"
    }, {
        "tihao": 57,
        "q": "以下关于字符串的叙述中，正确的是（ ）。",
        "a1": "A. 包含任意个空格字符的字符串称为空串",
        "a2": "B. 字符串不是线性数据结构",
        "a3": "C. 字符串的长度是指串中所含字符的个数",
        "a4": "D. 字符串的长度是指串中所含非空格字符的个数",
        "sl": "C"
    }, {
        "tihao": 58,
        "q": "已知栈S初始为空，用I表示入栈、O表示出栈，若入栈序列为a1a2a3a4a5，则通过栈S得到出栈序列a2a4a5a3a1的合法操作序列(  )。",
        "a1": "A. IIOIIOIOOO",
        "a2": "B. IOIOIOIOIO",
        "a3": "C. IOOIIOIOIO",
        "a4": "D. IIOOIOIOOO",
        "sl": "A"
    }, {
        "tihao": 59,
        "q": "某二叉树的先序遍历序列为ABCDEF ，中序遍历序列为BADCFE ，则该二叉树的高度(即层数)为( )。",
        "a1": "A. 3",
        "a2": "B. 4",
        "a3": "C. 5",
        "a4": "D. 6",
        "sl": "B"
    }, {
        "tihao": 60,
        "q": "对于n个元素的关键字序列{k1,k2, ...kn}，当且仅当满足关系ki≤k2i且ki≤k2i+1{i=1.2...[n/2]} 时称其为小根堆(小顶堆)。以下序列中，(   )不是小根堆。",
        "a1": "A. 16,25,40,55,30,50,45",
        "a2": "B. 16,40,25,50,45,30,55",
        "a3": "C. 16,25,39.,41,45,43,50",
        "a4": "D. 16,40,25,53,39,55,45",
        "sl": "D"
    }, {
        "tihao": 61,
        "q": "在12个互异元素构成的有序数组a[1..12] 中进行二分查找(即折半查找，向下取整)，若待查找的元素正好等于a[9]，则在此过程中，依次与数组中的( )比较后，查找成功结束。",
        "a1": "A. a[6]、a[7]、a[8]、a[9]",
        "a2": "B. a[6]、a[9]",
        "a3": "C. a[6]、a[7]、a[9]",
        "a4": "D. a[6]、a[8]、a[9]",
        "sl": "B"
    }, {
        "tihao": 62,
        "q": "某汽车加工工厂有两条装配线L1和L2，每条装配线的工位数均为n（Sij，i=1或2，j= 1，2，...，n），两条装配线对应的工位完成同样的加工工作，但是所需要的时间可能不同（aij，i=1或2，j=1，2，...，n）。汽车底盘开始到进入两条装配线的时间 (e1，e2) 以及装配后到结束的时间(x1x2)也可能不相同。从一个工位加工后流到下一个工位需要迁移时间(tij，i=1或2，j=2，...n）。现在要以最快的时间完成一辆汽车的装配，求最优的装配路线。分析该问题，发现问题具有最优子结构。以L1为例，除了第一个工位之外，经过第j个工位的最短时间包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，如式(1)。装配后到结束的最短时间包含离开L1的最短时间或者离开L2的最短时间如式（2）。由于在求解经过L1和L2的第j个工位的最短时间均包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，该问题具有重复子问题的性质，故采用迭代方法求解。该问题采用的算法设计策略是（62），算法的时间复杂度为（63）以下是一个装配调度实例，其最短的装配时间为（64），装配路线为（65）",
        "a1": "A. 分治",
        "a2": "B. 动态规划",
        "a3": "C. 贪心",
        "a4": "D. 回溯",
        "sl": "B"
    }, {
        "tihao": 63,
        "q": "某汽车加工工厂有两条装配线L1和L2，每条装配线的工位数均为n（Sij，i=1或2，j= 1，2，...，n），两条装配线对应的工位完成同样的加工工作，但是所需要的时间可能不同（aij，i=1或2，j=1，2，...，n）。汽车底盘开始到进入两条装配线的时间 (e1，e2) 以及装配后到结束的时间(x1x2)也可能不相同。从一个工位加工后流到下一个工位需要迁移时间(tij，i=1或2，j=2，...n）。现在要以最快的时间完成一辆汽车的装配，求最优的装配路线。分析该问题，发现问题具有最优子结构。以L1为例，除了第一个工位之外，经过第j个工位的最短时间包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，如式(1)。装配后到结束的最短时间包含离开L1的最短时间或者离开L2的最短时间如式（2）。由于在求解经过L1和L2的第j个工位的最短时间均包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，该问题具有重复子问题的性质，故采用迭代方法求解。该问题采用的算法设计策略是（62），算法的时间复杂度为（63）以下是一个装配调度实例，其最短的装配时间为（64），装配路线为（65）",
        "a1": "A. θ(lgn)",
        "a2": "B. θ(n)",
        "a3": "C. θ(n2)",
        "a4": "D. θ(nlgn)",
        "sl": "B"
    }, {
        "tihao": 64,
        "q": "某汽车加工工厂有两条装配线L1和L2，每条装配线的工位数均为n（Sij，i=1或2，j= 1，2，...，n），两条装配线对应的工位完成同样的加工工作，但是所需要的时间可能不同（aij，i=1或2，j=1，2，...，n）。汽车底盘开始到进入两条装配线的时间 (e1，e2) 以及装配后到结束的时间(x1x2)也可能不相同。从一个工位加工后流到下一个工位需要迁移时间(tij，i=1或2，j=2，...n）。现在要以最快的时间完成一辆汽车的装配，求最优的装配路线。分析该问题，发现问题具有最优子结构。以L1为例，除了第一个工位之外，经过第j个工位的最短时间包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，如式(1)。装配后到结束的最短时间包含离开L1的最短时间或者离开L2的最短时间如式（2）。由于在求解经过L1和L2的第j个工位的最短时间均包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，该问题具有重复子问题的性质，故采用迭代方法求解。该问题采用的算法设计策略是（62），算法的时间复杂度为（63）以下是一个装配调度实例，其最短的装配时间为（64），装配路线为（65）",
        "a1": "A. 21",
        "a2": "B. 23",
        "a3": "C. 20",
        "a4": "D. 26",
        "sl": "A"
    }, {
        "tihao": 65,
        "q": "某汽车加工工厂有两条装配线L1和L2，每条装配线的工位数均为n（Sij，i=1或2，j= 1，2，...，n），两条装配线对应的工位完成同样的加工工作，但是所需要的时间可能不同（aij，i=1或2，j=1，2，...，n）。汽车底盘开始到进入两条装配线的时间 (e1，e2) 以及装配后到结束的时间(x1x2)也可能不相同。从一个工位加工后流到下一个工位需要迁移时间(tij，i=1或2，j=2，...n）。现在要以最快的时间完成一辆汽车的装配，求最优的装配路线。分析该问题，发现问题具有最优子结构。以L1为例，除了第一个工位之外，经过第j个工位的最短时间包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，如式(1)。装配后到结束的最短时间包含离开L1的最短时间或者离开L2的最短时间如式（2）。由于在求解经过L1和L2的第j个工位的最短时间均包含了经过L1的第j-1个工位的最短时间或者经过L2的第j-1个工位的最短时间，该问题具有重复子问题的性质，故采用迭代方法求解。该问题采用的算法设计策略是（62），算法的时间复杂度为（63）以下是一个装配调度实例，其最短的装配时间为（64），装配路线为（65）",
        "a1": "A. S11→S12→S13",
        "a2": "B. S11→S22→S13",
        "a3": "C. S21→S12→S23",
        "a4": "D. S21→S22→S23",
        "sl": "B"
    }, {
        "tihao": 66,
        "q": "在浏览器地址栏输入一个正确的网址后，本地主机将首先在（）查询该网址对应的IP地址。",
        "a1": "A. 本地DNS缓存",
        "a2": "B. 本机hosts文件",
        "a3": "C. 本地DNS服务器",
        "a4": "D. 根域名服务器",
        "sl": "B"
    }, {
        "tihao": 67,
        "q": "下面关于Linux目录的描述中，正确的是（）。",
        "a1": "A. Linux只有一个根目录，用 ' / root '表示",
        "a2": "B. Linux中有多个根目录，用' / '加相应目录名称表示",
        "a3": "C. Linux中只有一个根目录，用' / '表示",
        "a4": "D. Linux中有多个根目录，用相应目录名称表示",
        "sl": "C"
    }, {
        "tihao": 68,
        "q": "以下关于TCP/IP协议栈中协议和层次的对应关系正确的是（）。",
        "a1": "A. ",
        "a2": "B. ",
        "a3": "C. ",
        "a4": "D. ",
        "sl": "C"
    }, {
        "tihao": 69,
        "q": "在异步通信中，每个字符包含1位起始位、7位数据位和2位终止位，若每秒钟传送500个字符，则有效数据速率为()。",
        "a1": "A. 500b/s",
        "a2": "B. 700b/s",
        "a3": "C. 3500b/s",
        "a4": "D. 5000b/s",
        "sl": "C"
    }, {
        "tihao": 70,
        "q": "以下路由策略中，依据网络信息经常更新路由的是( )。",
        "a1": "A. 静态路由",
        "a2": "B. 洪泛式",
        "a3": "C. 随机路由",
        "a4": "D. 自适应路由",
        "sl": "D"
    }, {
        "tihao": 71,
        "q": "The beauty of software is in its function，in its internal structure，and in the way in which it is created by a team. To a user，a program with just the right features presented through an intuitive and( 71 )interface is beautiful.To a software designer，an internal structure that is partitioned in a simple and intuitive manner，and that minimizes internal coupling is beautiful.To developers and managers ，a motivated team of developers making significant progress every week，and producing defect-free code，is beautiful.There is beauty on all these levels.our world needs software--lots of software. Fifty years ago software was something that ran in a few big and expensive machines. Thirty years ago it was something that ran in most companies and industrial settings. Now there is software running in our cell phones，watches，appliances，automobiles，toys，and tools. And need for new and better software never( 72 ).As our civilization grows and expands，as developing nations build their infrastructures，as developed nations strive to achieve ever greater efficiencies，the need for more and more Software( 73 )to increase. It would be a great shame if，in all that software，there was no beauty.We know that software can be ugly. We know that it can be hard to use，unreliable ，and carelessly structured. We know that there are software systems whose tangled and careless internal structures make them expensive and difficult to change. We know that there are software systems that present their features through an awkward and cumbersome interface. We know that there are software systems that crash and misbehave. These are( 74) systems. Unfortunately，as a profession，software developers tend to create more ugly systems than beautiful ones.There is a secret that the best software developers know. Beauty is cheaper than ugliness. Beauty is faster than ugliness. A beautiful software system can be built and maintained in less time，and for less money ,than an ugly one. Novice software developers don't. understand this. They think that they have to do everything fast and quick. They think that beauty is( 75 ) .No! By doing things fast and quick，they make messes that make the software stiff，and hard to understand，Beautiful systems e flexible and easy to understand. Building them and maintaining them is a joy. It is ugliness that is impractical.Ugliness will slow you down and make your software expensive and brittle. Beautiful systems cost the least build and maintain，and are delivered soonest.",
        "a1": "A. Simple",
        "a2": "B. Hard",
        "a3": "C. Complex",
        "a4": "D. duplicated",
        "sl": "A"
    }, {
        "tihao": 72,
        "q": "The beauty of software is in its function，in its internal structure，and in the way in which it is created by a team. To a user，a program with just the right features presented through an intuitive and( 71 )interface is beautiful.To a software designer，an internal structure that is partitioned in a simple and intuitive manner，and that minimizes internal coupling is beautiful.To developers and managers ，a motivated team of developers making significant progress every week，and producing defect-free code，is beautiful.There is beauty on all these levels.our world needs software--lots of software. Fifty years ago software was something that ran in a few big and expensive machines. Thirty years ago it was something that ran in most companies and industrial settings. Now there is software running in our cell phones，watches，appliances，automobiles，toys，and tools. And need for new and better software never( 72 ).As our civilization grows and expands，as developing nations build their infrastructures，as developed nations strive to achieve ever greater efficiencies，the need for more and more Software( 73 )to increase. It would be a great shame if，in all that software，there was no beauty.We know that software can be ugly. We know that it can be hard to use，unreliable ，and carelessly structured. We know that there are software systems whose tangled and careless internal structures make them expensive and difficult to change. We know that there are software systems that present their features through an awkward and cumbersome interface. We know that there are software systems that crash and misbehave. These are( 74) systems. Unfortunately，as a profession，software developers tend to create more ugly systems than beautiful ones.There is a secret that the best software developers know. Beauty is cheaper than ugliness. Beauty is faster than ugliness. A beautiful software system can be built and maintained in less time，and for less money ,than an ugly one. Novice software developers don't. understand this. They think that they have to do everything fast and quick. They think that beauty is( 75 ) .No! By doing things fast and quick，they make messes that make the software stiff，and hard to understand，Beautiful systems e flexible and easy to understand. Building them and maintaining them is a joy. It is ugliness that is impractical.Ugliness will slow you down and make your software expensive and brittle. Beautiful systems cost the least build and maintain，and are delivered soonest.",
        "a1": "A. happens",
        "a2": "B. exists",
        "a3": "C. stops",
        "a4": "D. starts",
        "sl": "C"
    }, {
        "tihao": 73,
        "q": "The beauty of software is in its function，in its internal structure，and in the way in which it is created by a team. To a user，a program with just the right features presented through an intuitive and( 71 )interface is beautiful.To a software designer，an internal structure that is partitioned in a simple and intuitive manner，and that minimizes internal coupling is beautiful.To developers and managers ，a motivated team of developers making significant progress every week，and producing defect-free code，is beautiful.There is beauty on all these levels.our world needs software--lots of software. Fifty years ago software was something that ran in a few big and expensive machines. Thirty years ago it was something that ran in most companies and industrial settings. Now there is software running in our cell phones，watches，appliances，automobiles，toys，and tools. And need for new and better software never( 72 ).As our civilization grows and expands，as developing nations build their infrastructures，as developed nations strive to achieve ever greater efficiencies，the need for more and more Software( 73 )to increase. It would be a great shame if，in all that software，there was no beauty.We know that software can be ugly. We know that it can be hard to use，unreliable ，and carelessly structured. We know that there are software systems whose tangled and careless internal structures make them expensive and difficult to change. We know that there are software systems that present their features through an awkward and cumbersome interface. We know that there are software systems that crash and misbehave. These are( 74) systems. Unfortunately，as a profession，software developers tend to create more ugly systems than beautiful ones.There is a secret that the best software developers know. Beauty is cheaper than ugliness. Beauty is faster than ugliness. A beautiful software system can be built and maintained in less time，and for less money ,than an ugly one. Novice software developers don't. understand this. They think that they have to do everything fast and quick. They think that beauty is( 75 ) .No! By doing things fast and quick，they make messes that make the software stiff，and hard to understand，Beautiful systems e flexible and easy to understand. Building them and maintaining them is a joy. It is ugliness that is impractical.Ugliness will slow you down and make your software expensive and brittle. Beautiful systems cost the least build and maintain，and are delivered soonest.",
        "a1": "A. starts",
        "a2": "B. continues",
        "a3": "C. appears",
        "a4": "D. stops",
        "sl": "B"
    }, {
        "tihao": 74,
        "q": "The beauty of software is in its function，in its internal structure，and in the way in which it is created by a team. To a user，a program with just the right features presented through an intuitive and( 71 )interface is beautiful.To a software designer，an internal structure that is partitioned in a simple and intuitive manner，and that minimizes internal coupling is beautiful.To developers and managers ，a motivated team of developers making significant progress every week，and producing defect-free code，is beautiful.There is beauty on all these levels.our world needs software--lots of software. Fifty years ago software was something that ran in a few big and expensive machines. Thirty years ago it was something that ran in most companies and industrial settings. Now there is software running in our cell phones，watches，appliances，automobiles，toys，and tools. And need for new and better software never( 72 ).As our civilization grows and expands，as developing nations build their infrastructures，as developed nations strive to achieve ever greater efficiencies，the need for more and more Software( 73 )to increase. It would be a great shame if，in all that software，there was no beauty.We know that software can be ugly. We know that it can be hard to use，unreliable ，and carelessly structured. We know that there are software systems whose tangled and careless internal structures make them expensive and difficult to change. We know that there are software systems that present their features through an awkward and cumbersome interface. We know that there are software systems that crash and misbehave. These are( 74) systems. Unfortunately，as a profession，software developers tend to create more ugly systems than beautiful ones.There is a secret that the best software developers know. Beauty is cheaper than ugliness. Beauty is faster than ugliness. A beautiful software system can be built and maintained in less time，and for less money ,than an ugly one. Novice software developers don't. understand this. They think that they have to do everything fast and quick. They think that beauty is( 75 ) .No! By doing things fast and quick，they make messes that make the software stiff，and hard to understand，Beautiful systems e flexible and easy to understand. Building them and maintaining them is a joy. It is ugliness that is impractical.Ugliness will slow you down and make your software expensive and brittle. Beautiful systems cost the least build and maintain，and are delivered soonest.",
        "a1": "A. practical",
        "a2": "B. useful",
        "a3": "C. beautiful",
        "a4": "D. ugly",
        "sl": "D"
    }, {
        "tihao": 75,
        "q": "The beauty of software is in its function，in its internal structure，and in the way in which it is created by a team. To a user，a program with just the right features presented through an intuitive and( 71 )interface is beautiful.To a software designer，an internal structure that is partitioned in a simple and intuitive manner，and that minimizes internal coupling is beautiful.To developers and managers ，a motivated team of developers making significant progress every week，and producing defect-free code，is beautiful.There is beauty on all these levels.our world needs software--lots of software. Fifty years ago software was something that ran in a few big and expensive machines. Thirty years ago it was something that ran in most companies and industrial settings. Now there is software running in our cell phones，watches，appliances，automobiles，toys，and tools. And need for new and better software never( 72 ).As our civilization grows and expands，as developing nations build their infrastructures，as developed nations strive to achieve ever greater efficiencies，the need for more and more Software( 73 )to increase. It would be a great shame if，in all that software，there was no beauty.We know that software can be ugly. We know that it can be hard to use，unreliable ，and carelessly structured. We know that there are software systems whose tangled and careless internal structures make them expensive and difficult to change. We know that there are software systems that present their features through an awkward and cumbersome interface. We know that there are software systems that crash and misbehave. These are( 74) systems. Unfortunately，as a profession，software developers tend to create more ugly systems than beautiful ones.There is a secret that the best software developers know. Beauty is cheaper than ugliness. Beauty is faster than ugliness. A beautiful software system can be built and maintained in less time，and for less money ,than an ugly one. Novice software developers don't. understand this. They think that they have to do everything fast and quick. They think that beauty is( 75 ) .No! By doing things fast and quick，they make messes that make the software stiff，and hard to understand，Beautiful systems e flexible and easy to understand. Building them and maintaining them is a joy. It is ugliness that is impractical.Ugliness will slow you down and make your software expensive and brittle. Beautiful systems cost the least build and maintain，and are delivered soonest.",
        "a1": "A. impractical",
        "a2": "B. perfect",
        "a3": "C. time-wasting",
        "a4": "D. practical",
        "sl": "A"
    }]
}
软件设计20年上={
	"code": 0,
	"list": [{
		"tihao": 1,
		"q": "在程序执行过程中，高速缓存(Cache) 与主存间的地址映射由（  ）。",
		"a1": "A. 程序员和操作系统共同协调完成",
		"a2": "B. 操作系统进行管理",
		"a3": "C. 程序员自行安排",
		"a4": "D. 硬件自动完成",
		"sl": "D"
	}, {
		"tihao": 2,
		"q": "计算机中提供指令地址的程序计数器PC在（  ）中。",
		"a1": "A. 控制器",
		"a2": "B. 运算器",
		"a3": "C. 存储器",
		"a4": "D. I/O设备",
		"sl": "A"
	}, {
		"tihao": 3,
		"q": "以下关于两个浮点数相加运算的叙述中，正确的是（  ）.",
		"a1": "A. 首先进行对阶，阶码大的向阶码小的对齐",
		"a2": "B. 首先进行对阶，阶码小的向阶码大的对齐",
		"a3": "C. 不需要对阶，直接将尾数相加",
		"a4": "D. 不需要对阶，直接将阶码相加",
		"sl": "B"
	}, {
		"tihao": 4,
		"q": "某计算机系统的CPU主频为2.8GHz。某应用程序包括3类指令，各类指令的CPI(执行每条指令所需要的时钟周期数)及指令比例如’下表所示。执行该应用程序时的平均CPI为（4）；运算速度用MIPS表示，约为（ 5）。",
		"a1": "A. 25",
		"a2": "B. 3",
		"a3": "C. 3.5",
		"a4": "D. 4",
		"sl": "C"
	}, {
		"tihao": 5,
		"q": "某计算机系统的CPU主频为2.8GHz。某应用程序包括3类指令，各类指令的CPI(执行每条指令所需要的时钟周期数)及指令比例如’下表所示。执行该应用程序时的平均CPI为（4）；运算速度用MIPS表示，约为（ 5）。",
		"a1": "A. 700",
		"a2": "B. 800",
		"a3": "C. 930",
		"a4": "D. 1100",
		"sl": "B"
	}, {
		"tihao": 6,
		"q": "中断向量提供（  ）.",
		"a1": "A. 函数调用结束后的返回地址",
		"a2": "B. I/O设备的接口地址",
		"a3": "C. 主程序的入口地址",
		"a4": "D. 中断服务程序入口地址",
		"sl": "D"
	}, {
		"tihao": 7,
		"q": "以下关于认证和加密的叙述中，错误的是（  ）。",
		"a1": "A. 加密用以确保数据的保密性",
		"a2": "B. 认证用以确保报文发送者和接收者的真实性",
		"a3": "C. 认证和加密都可以阻止对手进行被动攻击",
		"a4": "D. 身份认证的目的在于识别用户的合法性，阻止非法用户访问系统",
		"sl": "C"
	}, {
		"tihao": 8,
		"q": "访问控制是对信息系统资源进行保护的重要措施，适当的访问控制能够阻止未经授权的用户有意或者无意地获取资源。计算机系统中，访问控制的任务不包括（  ）。",
		"a1": "A. 审计",
		"a2": "B. 授权",
		"a3": "C. 确定存取权限",
		"a4": "D. 实施存取权限",
		"sl": "A"
	}, {
		"tihao": 9,
		"q": "路由协议称为内部网关协议，自治系统之间的协议称为外部网关协议，以下属于外部网关协议的是（  ）。",
		"a1": "A. RIP",
		"a2": "B. 0SPF",
		"a3": "C. BGP",
		"a4": "D. UDP",
		"sl": "C"
	}, {
		"tihao": 10,
		"q": "所有资源只能由授权方或以授权的方式进行修改，即信息未经授权不能进行改变的特性是指信息的（  ）。",
		"a1": "A. 完整性",
		"a2": "B. 可用性",
		"a3": "C. 保密性",
		"a4": "D. 不可抵赖性",
		"sl": "A"
	}, {
		"tihao": 11,
		"q": "在Windows操作系统下，要获取某个网络开放端口所对应的应用程序信息，可以使用命令（  ）。",
		"a1": "A. ipconfig",
		"a2": "B. traceroute",
		"a3": "C. netstat",
		"a4": "D. nslookup",
		"sl": "C"
	}, {
		"tihao": 12,
		"q": "甲、 乙两个申请人分别就相同内容的计算机软件发明创造，向国务院专利行政部门门提出专利申请，甲先于乙一日提出，则（  ）。",
		"a1": "A. 甲获得该项专利申请权",
		"a2": "B. 乙获得该项专利申请权",
		"a3": "C. 甲和乙都获得该项专利申请权",
		"a4": "D. 甲和乙都不能获得该项专利申请权",
		"sl": "A"
	}, {
		"tihao": 13,
		"q": "小王是某高校的非全8制在读研究生，目前在甲公司实习，负责了该公司某软件项目的开发工作并撰写相关的软件文档。以下叙述中，正确的是（  ）。",
		"a1": "A. 该软件文档属于职务作品，但小王享有该软件著作权的全部权利",
		"a2": "B. 该软件文档属于职务作品，甲公司享有该软件著作权的全部权利",
		"a3": "C. 该软件文档不属于职务作品，小王享有该软件著作权的全部权利",
		"a4": "D. 该软件文档不属于职务作品，甲公司和小王共同享有该著作权的全部权利",
		"sl": "B"
	}, {
		"tihao": 14,
		"q": "按照我国著作权法的权利保护期，以下权利中，（  ）受到永久保护。",
		"a1": "A. 发表权",
		"a2": "B. 修改权",
		"a3": "C. 复制权",
		"a4": "D. 发行权",
		"sl": "B"
	}, {
		"tihao": 15,
		"q": "结构化分析方法中，数据流图中的元素在（  ）中进行定义。",
		"a1": "A. 加工逻辑",
		"a2": "B. 实体联系图",
		"a3": "C. 流程图",
		"a4": "D. 数据字典",
		"sl": "D"
	}, {
		"tihao": 16,
		"q": "良好的启发式设计原则上不包括（  ）。",
		"a1": "A. 提高模块独立性",
		"a2": "B. 模块规模越小越好",
		"a3": "C. 模块作用域在其控制域之内",
		"a4": "D. 降低模块接口复杂性",
		"sl": "B"
	}, {
		"tihao": 17,
		"q": "如下所示的软件项目活动图中，顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的权重表示活动的持续时间(天)， 则完成该项目的最短时间为（17）天。在该活动图中，共有（18）条关键路径。 ",
		"a1": "A. 17",
		"a2": "B. 19",
		"a3": "C. 20",
		"a4": "D. 22",
		"sl": "D"
	}, {
		"tihao": 18,
		"q": "如下所示的软件项目活动图中，顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的权重表示活动的持续时间(天)， 则完成该项目的最短时间为（17）天。在该活动图中，共有（18）条关键路径。 ",
		"a1": "A. 1",
		"a2": "B. 2",
		"a3": "C. 3",
		"a4": "D. 4",
		"sl": "B"
	}, {
		"tihao": 19,
		"q": "软件项目成本估算模型COCOM01I中，体系结构阶段模型基于（  ）进行估算。",
		"a1": "A. 应用程序点数量",
		"a2": "B. 功能点数量",
		"a3": "C. 复用或生成的代码行数",
		"a4": "D. 源代码的行数",
		"sl": "D"
	}, {
		"tihao": 20,
		"q": "某表达式的语法树如下图所示，其后缀式(逆波兰式)是（  ）。",
		"a1": "A. abcd-+*",
		"a2": "B. ab-c+d*",
		"a3": "C. abc-d*+",
		"a4": "D. ab-cd+*",
		"sl": "C"
	}, {
		"tihao": 21,
		"q": "用C/C++语言为某个应用编写的程序，经过（  ）后形成可执行程序。",
		"a1": "A. 预处理、编译、汇编、链接",
		"a2": "B. 编译、预处理、汇编、链接",
		"a3": "C. 汇编、预处理、链接、编译",
		"a4": "D. 链接、预处理、编译、汇编",
		"sl": "A"
	}, {
		"tihao": 22,
		"q": "在程序的执行过程中，系统用（  ）实现嵌套调用(递归调用)函数的正确返回。",
		"a1": "A. 队列",
		"a2": "B. 优先队列",
		"a3": "C. 栈",
		"a4": "D. 散列表",
		"sl": "C"
	}, {
		"tihao": 23,
		"q": "假设系统中有三个进程P1、P2和P3，两种资源R1、R2。如果进程资源图如图①和图②所示，那么（  ）。",
		"a1": "A. 图①和图②都可化简",
		"a2": "B. 图①和图②都不可化简",
		"a3": "C. 图①可化简，图②不可化简",
		"a4": "D. 图①不可化简，图②可化简",
		"sl": "C"
	}, {
		"tihao": 24,
		"q": "假设计算机系统的页面大小为4K，进程P的页面变换表如下表所示。若P要动问的逻辑地址为十六进制3C20H，那么该逻辑地址经过地址变换后，其物理地址应为（  ）。",
		"a1": "A. 2048H",
		"a2": "B. 3C20H",
		"a3": "C. 5C20H",
		"a4": "D. 6C20H",
		"sl": "D"
	}, {
		"tihao": 25,
		"q": "某文件系统采用索引节点管理，其磁盘索引块和磁盘数据块大小均为1KB字节且每个文件索引节点有8个地址项iaddr[0]~iaddr[7]，每个地址项大小为4字节，其中iaddr[0]~iaddr[4]采用直接地址索引，iaddr[5]和iaddr[6]采用一级间接地址索引，iaddr[7] 采用二级间接地址索引。若用户要访问文件userA中逻辑块号为4和5的信息，则系统应分别采用（25）， 该文件系统可表示的单个文件最大长度是（ 26）KB。",
		"a1": "A. 直接地址访问和直接地址访问",
		"a2": "B. 直接地址访问和一级间接地址访问",
		"a3": "C. 一级问接地址访问和一级间接地址访问",
		"a4": "D. 一级间接地址访问和二级间接地址访问",
		"sl": "B"
	}, {
		"tihao": 26,
		"q": "某文件系统采用索引节点管理，其磁盘索引块和磁盘数据块大小均为1KB字节且每个文件索引节点有8个地址项iaddr[0]~iaddr[7]，每个地址项大小为4字节，其中iaddr[0]~iaddr[4]采用直接地址索引，iaddr[5]和iaddr[6]采用一级间接地址索引，iaddr[7] 采用二级间接地址索引。若用户要访问文件userA中逻辑块号为4和5的信息，则系统应分别采用（25）， 该文件系统可表示的单个文件最大长度是（ 26）KB。",
		"a1": "A. 517",
		"a2": "B. 1029",
		"a3": "C. 65797",
		"a4": "D. 66053",
		"sl": "D"
	}, {
		"tihao": 27,
		"q": "假设系统有n (n≥5) 个进程共享资源R，且资源R的可用数为5。若采用PV操作，则相应的信号量S的取值范围应为（  ）。",
		"a1": "A. -1~n-1",
		"a2": "B. -5~5",
		"a3": "C. -(n-1)~1",
		"a4": "D. -(n-5)~5",
		"sl": "D"
	}, {
		"tihao": 28,
		"q": "在支持多线程的操作系统中，假设进程P创建了线程T1、T2和T3， 那么以下叙述中错误的是（  ）。",
		"a1": "A. 线程T1、 T2和T3可以共享程P的代码",
		"a2": "B. 线程T1、T2可以共享P进程中T3的栈指针",
		"a3": "C. 线程T1、T2和T3可以共享进程P打开的文件",
		"a4": "D. 线程T1、T2和T3可以共享进程P的全局变量",
		"sl": "B"
	}, {
		"tihao": 29,
		"q": "喷泉模型是一种适合于面向（29）开发方法的软件过程模型。该过程模型的特点不包括（30）。",
		"a1": "A. 对象",
		"a2": "B. 数据",
		"a3": "C. 数据流",
		"a4": "D. 事件",
		"sl": "A"
	}, {
		"tihao": 30,
		"q": "喷泉模型是一种适合于面向（29）开发方法的软件过程模型。该过程模型的特点不包括（30）。",
		"a1": "A. 以用户需求为动力",
		"a2": "B. 支持软件重用",
		"a3": "C. 具有迭代性",
		"a4": "D. 开发活动之间存在明显的界限",
		"sl": "D"
	}, {
		"tihao": 31,
		"q": "若某模块内所有处理元素都在同一个数据结构上操作，则该模块的内聚类型为（  ）。",
		"a1": "A. 逻辑",
		"a2": "B. 过程",
		"a3": "C. 通信",
		"a4": "D. 功能",
		"sl": "C"
	}, {
		"tihao": 32,
		"q": "软件质量属性中，（  ）是指软件每分钟可以处理多少个请求。",
		"a1": "A. 响应时间",
		"a2": "B. 吞吐量",
		"a3": "C. 负载",
		"a4": "D. 容量",
		"sl": "B"
	}, {
		"tihao": 33,
		"q": "提高程序执行效率的方法一般不包括（  ）。",
		"a1": "A. 设计更好的算法",
		"a2": "B. 采用不同的数据结构",
		"a3": "C. 采用不同的程序设计语言",
		"a4": "D. 改写代码使其更紧凑",
		"sl": "D"
	}, {
		"tihao": 34,
		"q": "软件可靠性是指系统在给定的时间间隔内、在给定条件下无失效运行的概率。若MTTF和MTTR分别表示平均无故障时间和平均修复时间，则公式（  ）可用于计算软件可靠性。",
		"a1": "A. MTTF/(1+MTTF)",
		"a2": "B. 1/(1+MTTF)",
		"a3": "C. MTTR/(1+MTTR)",
		"a4": "D. 1/(1+MTTR)",
		"sl": "A"
	}, {
		"tihao": 35,
		"q": "用白盒测试技术对下面流程图进行测试，设计的测试用例如下表所示。至少采用测试用例（35）才可以实现语句覆盖；至少采用测试用例（36）才可以实现路径覆盖。",
		"a1": "A. ①",
		"a2": "B. ②",
		"a3": "C. ③",
		"a4": "D. ④",
		"sl": "A"
	}, {
		"tihao": 36,
		"q": "用白盒测试技术对下面流程图进行测试，设计的测试用例如下表所示。至少采用测试用例（35）才可以实现语句覆盖；至少采用测试用例（36）才可以实现路径覆盖。",
		"a1": "A. ①",
		"a2": "B. ①②",
		"a3": "C. ③④",
		"a4": "D. ①②③④",
		"sl": "D"
	}, {
		"tihao": 37,
		"q": "面向对象程序设计语言C++、JAVA中，关键字（  ）可以用于区分同名的对象属性和局部变量名。",
		"a1": "A. private",
		"a2": "B. protected",
		"a3": "C. public",
		"a4": "D. this",
		"sl": "D"
	}, {
		"tihao": 38,
		"q": "采用面向对象方法进行系统开发时，以下与新型冠状病毒有关的对象中，存在“一般-特殊’关系的是（  ）。",
		"a1": "A. 确诊病人和治愈病人",
		"a2": "B. 确诊病人和疑似病人",
		"a3": "C. 医生和病人",
		"a4": "D. 发热病人和确诊病人",
		"sl": "D"
	}, {
		"tihao": 39,
		"q": "进行面向对象系统设计时，针对包中的所有类对于同-类性质的变化；一个变化若对一个包产生影响，则将对该包中的所有类产生影响，而对于其他的包不造成任何影响。这属于（  ）设计原则。",
		"a1": "A. 共同重用",
		"a2": "B. 开放-封闭",
		"a3": "C. 接口分离",
		"a4": "D. 共同封闭",
		"sl": "D"
	}, {
		"tihao": 40,
		"q": "多态有不同的形式，（  ）的多态是指同一个名字在不同上下文中所代表的含义不同。",
		"a1": "A. 参数",
		"a2": "B. 包含",
		"a3": "C. 过载",
		"a4": "D. 强制",
		"sl": "C"
	}, {
		"tihao": 41,
		"q": "某类图如图所示，下列选项错误的是（  ）。 ",
		"a1": "A. 一个A1的对象可能与一个A2的对象关联",
		"a2": "B. 一个A的非直接对象可能与一个A1的对象关联",
		"a3": "C. 类B1的对象可能通过A2与C1的对象关联",
		"a4": "D. 有可能A的直接对象与B1的对象关联",
		"sl": "D"
	}, {
		"tihao": 42,
		"q": "UML图中，对象图展现了（42），（43）所示对象图与 下图所示类图不一致。 ",
		"a1": "A. 一组对象、接口、协作和它们之间的关系",
		"a2": "B. 一组用例、参与者以及它们之间的关系",
		"a3": "C. 某一时刻一组对象以及它们之间的关系",
		"a4": "D. 以时间顺序组织的对象之间的交互活动",
		"sl": "C"
	}, {
		"tihao": 43,
		"q": "UML图中，对象图展现了（42），（43）所示对象图与 下图所示类图不一致。 ",
		"a1": "A. ",
		"a2": "B. ",
		"a3": "C. ",
		"a4": "D. ",
		"sl": "D"
	}, {
		"tihao": 44,
		"q": "某快餐厅主要制作并出售儿童套餐，一般包括主餐(各类比萨)、饮料和玩具，其餐品种类可能不同，但制作过程相同。前台服务员(Waiter) 调度厨师制作套餐。欲开发一软件，实现该制作过程，设计如下所示类图。该设计采用（44）模式将一个复杂对象的构建与它的表示分离，使得同样的构建过程可以创建不同的表示。其中，（45）构造一个使用Builder接口的对象。该模式属于（46）模式，该模式适用于（47）的情况。",
		"a1": "A. 生成器(Builder)",
		"a2": "B. 抽象工厂(Abstract Factory)",
		"a3": "C. 原型(Prototype)",
		"a4": "D. 工厂方法(Factory Method)",
		"sl": "A"
	}, {
		"tihao": 45,
		"q": "某快餐厅主要制作并出售儿童套餐，一般包括主餐(各类比萨)、饮料和玩具，其餐品种类可能不同，但制作过程相同。前台服务员(Waiter) 调度厨师制作套餐。欲开发一软件，实现该制作过程，设计如下所示类图。该设计采用（44）模式将一个复杂对象的构建与它的表示分离，使得同样的构建过程可以创建不同的表示。其中，（45）构造一个使用Builder接口的对象。该模式属于（46）模式，该模式适用于（47）的情况。",
		"a1": "A. PizzaBuilder",
		"a2": "B. SpicyPizaBuilder",
		"a3": "C. Waiter",
		"a4": "D. Pizza",
		"sl": "C"
	}, {
		"tihao": 46,
		"q": "某快餐厅主要制作并出售儿童套餐，一般包括主餐(各类比萨)、饮料和玩具，其餐品种类可能不同，但制作过程相同。前台服务员(Waiter) 调度厨师制作套餐。欲开发一软件，实现该制作过程，设计如下所示类图。该设计采用（44）模式将一个复杂对象的构建与它的表示分离，使得同样的构建过程可以创建不同的表示。其中，（45）构造一个使用Builder接口的对象。该模式属于（46）模式，该模式适用于（47）的情况。",
		"a1": "A. 创建型对象",
		"a2": "B. 结构型对象",
		"a3": "C. 行为型对象",
		"a4": "D. 结构型类",
		"sl": "A"
	}, {
		"tihao": 47,
		"q": "某快餐厅主要制作并出售儿童套餐，一般包括主餐(各类比萨)、饮料和玩具，其餐品种类可能不同，但制作过程相同。前台服务员(Waiter) 调度厨师制作套餐。欲开发一软件，实现该制作过程，设计如下所示类图。该设计采用（44）模式将一个复杂对象的构建与它的表示分离，使得同样的构建过程可以创建不同的表示。其中，（45）构造一个使用Builder接口的对象。该模式属于（46）模式，该模式适用于（47）的情况。",
		"a1": "A. 当一个系统应该独立于它的产品创建、构成和表示时",
		"a2": "B. 当一个类希望由它的子类来指定它所创建的对象的时候",
		"a3": "C. 当要强调一系列相关的产品对象的设计以便进行联合使用时",
		"a4": "D. 当构造过程必须允许被构造的对象有不同的表示时",
		"sl": "D"
	}, {
		"tihao": 48,
		"q": "函数foo()、hoo0定义如下，调用函数hoo()时，第-个参数采用传值(call by value)方式，第二个参数采用传引用(call by reference)方式。设有函数调(函数foo(5)，那么\"print(x)”执行后输出的值为（  ）。",
		"a1": "A. 24",
		"a2": "B. 25",
		"a3": "C. 30",
		"a4": "D. 36",
		"sl": "A"
	}, {
		"tihao": 49,
		"q": "程序设计语言的大多数语法现象可以用CFG (上 下文无关文法)表示。下面的CFG产生式集用于描述简单算术表达式，其中+、-、*表示加、减、乘运算，id表示单个字母表示的变量，那么符合该文法的表达式为（  ）。P:E→E+T|E-T|TT→T*F|FF→F|id",
		"a1": "A. a+-b-c",
		"a2": "B. a*(b+c)",
		"a3": "C. a*-b+2",
		"a4": "D. -a/b+c",
		"sl": "A"
	}, {
		"tihao": 50,
		"q": "某有限自动机的状态转换图如下图所示，该自动机可识别（  ）。",
		"a1": "A. 1001",
		"a2": "B. 1100",
		"a3": "C. 1010",
		"a4": "D. 0101",
		"sl": "B"
	}, {
		"tihao": 51,
		"q": "某高校信息系统设计的分E-R图中，人力部门定义的职工实体具有属性：职工号、姓名、性别和出生日期；教学部门定义的教师实体具有属性：教师号、姓名和职称。这种情况属于（51），在合并E-R图时，（52）解决这一冲突。",
		"a1": "A. 属性冲突",
		"a2": "B. 命名冲突",
		"a3": "C. 结构冲突",
		"a4": "D. 实体冲突",
		"sl": "C"
	}, {
		"tihao": 52,
		"q": "某高校信息系统设计的分E-R图中，人力部门定义的职工实体具有属性：职工号、姓名、性别和出生日期；教学部门定义的教师实体具有属性：教师号、姓名和职称。这种情况属于（51），在合并E-R图时，（52）解决这一冲突。",
		"a1": "A. 职工和教师实体保持各自属性不变",
		"a2": "B. 职工实体中加入职称属性，删除教师实体",
		"a3": "C. 教师也是学校的职工，故直接将教师实体删除",
		"a4": "D. 将教师实体所有属性并入职工实体，删除教师实体",
		"sl": "B"
	}, {
		"tihao": 53,
		"q": "假设关系R, U={A,B,C,D,E}, F= {A→BC,AC→D,B→D}，那么在关系R中（  ）。",
		"a1": "A. 不存在传递依赖，候选关键字A",
		"a2": "B. 不存在传递依赖，候选关键字AC",
		"a3": "C. 存在传递依赖A→D，候选关键字A",
		"a4": "D. 存在传递依赖B→D，候选关键字C",
		"sl": "C"
	}, {
		"tihao": 54,
		"q": "关系R、S如下表所示，的结果集为（54），R、S的左外联接、右外联接和完全外联接的元组个数分别为（55）。",
		"a1": "A. { (2,1,4),(3,4,4)}",
		"a2": "B. { (2,1,4,8),(3,4,4,4)}",
		"a3": "C. { (C,1.4.2,1.8).(3.4.4.3,4,4)}",
		"a4": "D. { (1,2,3,1,9,1),(2,1,4,2,1,8),(3,4,4,3,4,4).(4,6,7.4,8,3)}",
		"sl": "B"
	}, {
		"tihao": 55,
		"q": "关系R、S如下表所示，的结果集为（54），R、S的左外联接、右外联接和完全外联接的元组个数分别为（55）。",
		"a1": "A. 2,2,4",
		"a2": "B. 2,2,6",
		"a3": "C. 4,4,4",
		"a4": "D. 4,4,6",
		"sl": "D"
	}, {
		"tihao": 56,
		"q": "某企业信息系统采用分布式数据库系统。”当某一场地故障时，系统可以使用其他场地上的副本而不至于使整个系统瘫痪\"称为分布式数据库的（  ）。",
		"a1": "A. 共享性",
		"a2": "B. 自治性",
		"a3": "C. 可用性",
		"a4": "D. 分布性",
		"sl": "C"
	}, {
		"tihao": 57,
		"q": "以下关于Huffman (哈夫曼)树的叙述中，错误的是（  ）。",
		"a1": "A. 权值越大的叶子离根结点越近",
		"a2": "B. Huffman (哈夫曼)树中不存在只有一个子树的结点",
		"a3": "C. Huffman (哈夫曼)树中的结点总数一定为奇数",
		"a4": "D. 权值相同的结点到树根的路径长度一定相同",
		"sl": "D"
	}, {
		"tihao": 58,
		"q": "通过元素在存储空间中的相对位置来表示数据元素之间的逻辑关系，是（  ）的特点。",
		"a1": "A. 顺序存储",
		"a2": "B. 链表存储",
		"a3": "C. 索引存储",
		"a4": "D. 哈希存储",
		"sl": "A"
	}, {
		"tihao": 59,
		"q": "在线性表L中进行二分查找，要求L（  ）。",
		"a1": "A. 顺序存储，元素随机排列",
		"a2": "B. 双向链表存储，元素随机排列",
		"a3": "C. 顺序存储，元素有序排列",
		"a4": "D. 双向链表存储，元素有序排列",
		"sl": "C"
	}, {
		"tihao": 60,
		"q": "某有向图如下所示，从顶点v1出发对其进行深度优先遍历，可能能得到的遍历序列是（60）; 从顶点v1出发对其进行广度优先遍历，可能得到的遍历序列是（61）。 ①v1 v2v3 v4 v5②v1 v3 v4v5v2③v1 v3v2v4 v5④v1 v2v4v5 v3",
		"a1": "A. ①②③",
		"a2": "B. ①③④",
		"a3": "C. ①②④",
		"a4": "D. ②③④",
		"sl": "D"
	}, {
		"tihao": 61,
		"q": "某有向图如下所示，从顶点v1出发对其进行深度优先遍历，可能能得到的遍历序列是（60）; 从顶点v1出发对其进行广度优先遍历，可能得到的遍历序列是（61）。 ①v1 v2v3 v4 v5②v1 v3 v4v5v2③v1 v3v2v4 v5④v1 v2v4v5 v3",
		"a1": "A. ①②",
		"a2": "B. ①③",
		"a3": "C. ②③",
		"a4": "D. ③④",
		"sl": "B"
	}, {
		"tihao": 62,
		"q": "对数组A=(2,8,7,1,3,5,6,4)用快速排序算法的划分方法进行一趟划分后得到的数组A为（62）(非递减排序， 以最后一个元素为基准元素)。进行一趟划分的计算时间为（63）。",
		"a1": "A. (1,2,8,7,3,5,6,4)",
		"a2": "B. (1,2,3,4,8,7,5,6)",
		"a3": "C. (2,3,1,4,7,5,6,8)",
		"a4": "D. (2,1,3,4,8,7,5,6)",
		"sl": "C"
	}, {
		"tihao": 63,
		"q": "对数组A=(2,8,7,1,3,5,6,4)用快速排序算法的划分方法进行一趟划分后得到的数组A为（62）(非递减排序， 以最后一个元素为基准元素)。进行一趟划分的计算时间为（63）。",
		"a1": "A. O(1)",
		"a2": "B. O(Ign)",
		"a3": "C. O(n)",
		"a4": "D. O(nlgn)",
		"sl": "C"
	}, {
		"tihao": 64,
		"q": "某简单无向连通图G的顶点数为n，则图G最少和最多分别有（  ）条边。",
		"a1": "A. n,n2/2",
		"a2": "B. n-1,n*(n-1)/2",
		"a3": "C. n,n*(n-1)/2",
		"a4": "D. n-1,n2/2",
		"sl": "B"
	}, {
		"tihao": 65,
		"q": "根据渐进分析，表达式序列：n4, lgn, 2n, 1000n, n2/3, n!从低到高排序为（  ）。",
		"a1": "A. Ign,1000n, n2/3, n4, n!, 2n",
		"a2": "B. n2/3,1000n, lgn, n4, n!, 2n",
		"a3": "C. lgn,1000n, n2/3, 2n, n4, n!",
		"a4": "D. Ign, n2/3, 1000n, n4, 2n, n!",
		"sl": "D"
	}, {
		"tihao": 66,
		"q": "采用DHCP动态分配IP地址，如果某主机开机后没有得到DHCP服务器的响应。则该主机获取的IP地址属于网络（  ）。",
		"a1": "A. 202.117.0.0/24",
		"a2": "B. 192.168.1.0/24",
		"a3": "C. 172.16.0.0/16",
		"a4": "D. 169.254.0.0/16",
		"sl": "D"
	}, {
		"tihao": 67,
		"q": "在浏览器的地址栏中输入xxxyftp.abc.can.cn，在该URL中（ ）是要访问的主机名。",
		"a1": "A. xxxftp",
		"a2": "B. abc",
		"a3": "C. can",
		"a4": "D. cn",
		"sl": "A"
	}, {
		"tihao": 68,
		"q": "当修改邮件时，客户与POP3服务器之间通过（68）建立连接，所使用的端口是（69）。",
		"a1": "A. HTTP",
		"a2": "B. TCP",
		"a3": "C. UDP",
		"a4": "D. HTTPS",
		"sl": "B"
	}, {
		"tihao": 69,
		"q": "当修改邮件时，客户与POP3服务器之间通过（68）建立连接，所使用的端口是（69）。",
		"a1": "A. 52",
		"a2": "B. 25",
		"a3": "C. 1100",
		"a4": "D. 110",
		"sl": "D"
	}, {
		"tihao": 70,
		"q": "因特网中的域名系统(Domain Name System)是一个分层的域名，在根域下面是顶级域，以下顶级域中，（  ）属于国家顶级域。",
		"a1": "A. NET",
		"a2": "B. EDU",
		"a3": "C. COM",
		"a4": "D. UK",
		"sl": "D"
	}, {
		"tihao": 71,
		"q": "Regardless of how well designed, constructed, and tested a system or application may be, errors or bugs will inevitably occur. Once a system has been（71），it enters operations and support.Systems support is the ongoing technical support for user, as well as the maintenance required to fix any errors, omissions,or new requirements that may arise. Before an information system can be（72）, it must be in operation. System operation is the day-to-day, week-to-week, month-to-month, and year-t-year（73）of an information system's business processes and application programs.Unlike systems analysis, design, and implementation, systems support cannot sensibly be（74）into actual phases that a support project must perform. Rather, systems support consists of four ongoing activities that are program maintenance, system recovery, technical support, and system enhancement.Each activity is a type of support project that is（75）by a particular problem,event, or opportunity encountered with the implemented system.",
		"a1": "A. designed",
		"a2": "B. implemented",
		"a3": "C. investigated",
		"a4": "D. analyzed",
		"sl": "B"
	}, {
		"tihao": 72,
		"q": "Regardless of how well designed, constructed, and tested a system or application may be, errors or bugs will inevitably occur. Once a system has been（71），it enters operations and support.Systems support is the ongoing technical support for user, as well as the maintenance required to fix any errors, omissions,or new requirements that may arise. Before an information system can be（72）, it must be in operation. System operation is the day-to-day, week-to-week, month-to-month, and year-t-year（73）of an information system's business processes and application programs.Unlike systems analysis, design, and implementation, systems support cannot sensibly be（74）into actual phases that a support project must perform. Rather, systems support consists of four ongoing activities that are program maintenance, system recovery, technical support, and system enhancement.Each activity is a type of support project that is（75）by a particular problem,event, or opportunity encountered with the implemented system.",
		"a1": "A. supported",
		"a2": "B. tested",
		"a3": "C. implemented",
		"a4": "D. constructed",
		"sl": "A"
	}, {
		"tihao": 73,
		"q": "Regardless of how well designed, constructed, and tested a system or application may be, errors or bugs will inevitably occur. Once a system has been（71），it enters operations and support.Systems support is the ongoing technical support for user, as well as the maintenance required to fix any errors, omissions,or new requirements that may arise. Before an information system can be（72）, it must be in operation. System operation is the day-to-day, week-to-week, month-to-month, and year-t-year（73）of an information system's business processes and application programs.Unlike systems analysis, design, and implementation, systems support cannot sensibly be（74）into actual phases that a support project must perform. Rather, systems support consists of four ongoing activities that are program maintenance, system recovery, technical support, and system enhancement.Each activity is a type of support project that is（75）by a particular problem,event, or opportunity encountered with the implemented system.",
		"a1": "A. construction",
		"a2": "B. maintenance",
		"a3": "C. execution",
		"a4": "D. implementation",
		"sl": "C"
	}, {
		"tihao": 74,
		"q": "Regardless of how well designed, constructed, and tested a system or application may be, errors or bugs will inevitably occur. Once a system has been（71），it enters operations and support.Systems support is the ongoing technical support for user, as well as the maintenance required to fix any errors, omissions,or new requirements that may arise. Before an information system can be（72）, it must be in operation. System operation is the day-to-day, week-to-week, month-to-month, and year-t-year（73）of an information system's business processes and application programs.Unlike systems analysis, design, and implementation, systems support cannot sensibly be（74）into actual phases that a support project must perform. Rather, systems support consists of four ongoing activities that are program maintenance, system recovery, technical support, and system enhancement.Each activity is a type of support project that is（75）by a particular problem,event, or opportunity encountered with the implemented system.",
		"a1": "A. broke",
		"a2": "B. formed",
		"a3": "C. composed",
		"a4": "D. decomposed",
		"sl": "D"
	}, {
		"tihao": 75,
		"q": "Regardless of how well designed, constructed, and tested a system or application may be, errors or bugs will inevitably occur. Once a system has been（71），it enters operations and support.Systems support is the ongoing technical support for user, as well as the maintenance required to fix any errors, omissions,or new requirements that may arise. Before an information system can be（72）, it must be in operation. System operation is the day-to-day, week-to-week, month-to-month, and year-t-year（73）of an information system's business processes and application programs.Unlike systems analysis, design, and implementation, systems support cannot sensibly be（74）into actual phases that a support project must perform. Rather, systems support consists of four ongoing activities that are program maintenance, system recovery, technical support, and system enhancement.Each activity is a type of support project that is（75）by a particular problem,event, or opportunity encountered with the implemented system.",
		"a1": "A. triggered",
		"a2": "B. leaded",
		"a3": "C. caused",
		"a4": "D. produced",
		"sl": "A"
	}]
}
软件设计16年上={
	"code": 0,
	"list": [{
		"tihao": 1,
		"q": "VLIW是（1）的简称。",
		"a1": "A. 复杂指令系统计算机",
		"a2": "B. 超大规模集成电路",
		"a3": "C. 单指令流多数据流",
		"a4": "D. 超长指令字",
		"sl": "D"
	}, {
		"tihao": 2,
		"q": "主存与Cache的地址映射方式中，（2）方式可以实现主存任意一块装入Cache中任意位置，只有装满才需要替换。",
		"a1": "A. 全相联",
		"a2": "B. 直接映射",
		"a3": "C. 组相联",
		"a4": "D. 串并联",
		"sl": "A"
	}, {
		"tihao": 3,
		"q": "如果《2x》的补码是《90H》，那么x的真值是（3）。",
		"a1": "A. 72",
		"a2": "B. -56",
		"a3": "C. 56",
		"a4": "D. 111",
		"sl": "B"
	}, {
		"tihao": 4,
		"q": "移位指令中的（4）指令的操作结果相当于对操作数进行乘2操作。",
		"a1": "A. 算术左移",
		"a2": "B. 逻辑右移",
		"a3": "C. 算术右移",
		"a4": "D. 带进位循环左移",
		"sl": "A"
	}, {
		"tihao": 5,
		"q": "内存按字节编址，从A1000H到B13FFH的区域的存储容量为（5）KB。",
		"a1": "A. 32",
		"a2": "B. 34",
		"a3": "C. 65",
		"a4": "D. 67",
		"sl": "C"
	}, {
		"tihao": 6,
		"q": "以下关于总线的叙述中，不正确的是（6）。",
		"a1": "A. 并行总线适合近距离高速数据传输",
		"a2": "B. 串行总线适合长距离数据传输",
		"a3": "C. 单总线结构在一个总线上适应不同种类的设备，设计简单且性能很高",
		"a4": "D. 专用总线在设计上可以与连接设备实现最佳匹配",
		"sl": "C"
	}, {
		"tihao": 7,
		"q": "以下关于网络层次与主要设备对应关系的叙述中，配对正确的是（7）。",
		"a1": "A. 网络层——集线器",
		"a2": "B. 数据链路层——网桥",
		"a3": "C. 传输层——路由器",
		"a4": "D. 会话层——防火墙",
		"sl": "B"
	}, {
		"tihao": 8,
		"q": "传输经过SSL加密的网页所采用的协议是（8）。",
		"a1": "A. HTTP",
		"a2": "B. HTTPS",
		"a3": "C. S-HTTP",
		"a4": "D. HTTP-S",
		"sl": "B"
	}, {
		"tihao": 9,
		"q": "为了攻击远程主机，通常利用（9）技术检测远程主机状态。",
		"a1": "A. 病毒查杀",
		"a2": "B. 端口扫描",
		"a3": "C. QQ聊天",
		"a4": "D. 身份认证",
		"sl": "B"
	}, {
		"tihao": 10,
		"q": "某软件公司参与开发管理系统软件的程序员张某，辞职到另一公司任职，于是该项目负责人将该管理系统软件上开发者的署名更改为李某（接张某工作）。该项目负责人的行为（10）。",
		"a1": "A. 侵犯了张某开发者身份权（署名权）",
		"a2": "B. 不构成侵权，因为程序员张某不是软件著作权人",
		"a3": "C. 只是行使管理者的权利，不构成侵权",
		"a4": "D. 不构成侵权，因为程序员张某现已不是项目组成员",
		"sl": "A"
	}, {
		"tihao": 11,
		"q": "美国某公司与中国某企业谈技术合作，合同约定使用l项美国专利（获得批准并在有效期内），该项技术未在中国和其他国家申请专利。依照该专利生产的产品（11）需要向美国公司支付这件美国专利的许可使用费。",
		"a1": "A. 在中国销售，中国企业",
		"a2": "B. 如果返销美国，中国企业不",
		"a3": "C. 在其他国家销售，中国企业",
		"a4": "D. 在中国销售，中国企业不",
		"sl": "D"
	}, {
		"tihao": 12,
		"q": "以下媒体文件格式中，（12）是视频文件格式。",
		"a1": "A. WAV",
		"a2": "B. BMP",
		"a3": "C. MP3",
		"a4": "D. MOV",
		"sl": "D"
	}, {
		"tihao": 13,
		"q": "以下软件产品中，属于图像编辑处理工具的软件是（13）。",
		"a1": "A. Powerpoint",
		"a2": "B. Photoshop",
		"a3": "C. Premiere",
		"a4": "D. Acrobat",
		"sl": "B"
	}, {
		"tihao": 14,
		"q": "使用150DPI的扫描分辨率扫描一幅3×4英寸的彩色照片，得到原始的24位真彩色图像的数据量是（14）Byte。",
		"a1": "A. 1800",
		"a2": "B. 90000",
		"a3": "C. 270000",
		"a4": "D. 810000",
		"sl": "D"
	}, {
		"tihao": 15,
		"q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示活动的持续时间（天），则完成该项目的最少时间为（15）天。活动BD最多可以晚开始（16）天而不会影响整个项目的进度。",
		"a1": "A. 15",
		"a2": "B. 21",
		"a3": "C. 22",
		"a4": "D. 24",
		"sl": "C"
	}, {
		"tihao": 16,
		"q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示活动的持续时间（天），则完成该项目的最少时间为（15）天。活动BD最多可以晚开始（16）天而不会影响整个项目的进度。",
		"a1": "A. 0",
		"a2": "B. 2",
		"a3": "C. 3",
		"a4": "D. 5",
		"sl": "A"
	}, {
		"tihao": 17,
		"q": "在结构化分析中，用数据流图描述（17）。当采用数据流图对一个图书馆管理系统进行分析时，（18）是一个外部实体。",
		"a1": "A. 数据对象之间的关系，用于对数据建模",
		"a2": "B. 数据在系统中如何被传送或变换，以及如何对数据流进行变换的功能或子功能，用于对功能建模",
		"a3": "C. 系统对外部事件如何响应，如何动作，用于对行为建模",
		"a4": "D. 数据流图中的各个组成部分",
		"sl": "B"
	}, {
		"tihao": 18,
		"q": "在结构化分析中，用数据流图描述（17）。当采用数据流图对一个图书馆管理系统进行分析时，（18）是一个外部实体。",
		"a1": "A. 读者",
		"a2": "B. 图书",
		"a3": "C. 借书证",
		"a4": "D. 借阅",
		"sl": "A"
	}, {
		"tihao": 19,
		"q": "软件开发过程中，需求分析阶段的输出不包括（19）。",
		"a1": "A. 流图",
		"a2": "B. 实体联系图",
		"a3": "C. 数据字典",
		"a4": "D. 软件体系结构图",
		"sl": "D"
	}, {
		"tihao": 20,
		"q": "以下关于高级程序设计语言实现的编译和解释方式的叙述中，正确的是（20）。",
		"a1": "A. 编译程序不参与用户程序的运行控制，而解释程序则参与",
		"a2": "B. 编译程序可以用高级语言编写，而解释程序只能用汇编语言编写",
		"a3": "C. 编译方式处理源程序时不进行优化，而解释方式则进行优化",
		"a4": "D. 编译方式不生成源程序的目标程序，而解释方式则生成",
		"sl": "A"
	}, {
		"tihao": 21,
		"q": "以下关于脚本语言的叙述中，正确的是（21）。",
		"a1": "A. 脚本语言是通用的程序设计语言",
		"a2": "B. 脚本语言更适合应用在系统级程序开发中",
		"a3": "C. 脚本语言主要采用解释方式实现",
		"a4": "D. 脚本语言中不能定义函数和调用函数",
		"sl": "C"
	}, {
		"tihao": 22,
		"q": "将高级语言源程序先转化为一种中间代码是现代编译器的常见处理方式。常用的中间代码有后缀式、（22）、语法树等。",
		"a1": "A. 前缀码",
		"a2": "B. 三地址码",
		"a3": "C. 符号表",
		"a4": "D. 补码和移码",
		"sl": "B"
	}, {
		"tihao": 23,
		"q": "当用户通过键盘或鼠标进入某应用系统时，通常最先获得键盘或鼠标输入信息的是（23）程序。",
		"a1": "A. 命令解释",
		"a2": "B. 中断处理",
		"a3": "C. 用户登录",
		"a4": "D. 系统调用",
		"sl": "B"
	}, {
		"tihao": 24,
		"q": "在Windows操作系统中，当用户双击《IMG_20160122_103.jpg》文件名时，系统会自动通过建立的（24）来决定使用什么程序打开该图像文件。",
		"a1": "A. 文件",
		"a2": "B. 文件关联",
		"a3": "C. 文件目录",
		"a4": "D. 临时文件",
		"sl": "B"
	}, {
		"tihao": 25,
		"q": "某磁盘有100个磁道，磁头从一个磁道移至另一个磁道需要6ms。文件在磁盘上非连续存放，逻辑上相邻数据块的平均距离为10个磁道，每块的旋转延迟时间及传输时间分别为100ms和20ms，则读取一个100块的文件需要（25）ms。",
		"a1": "A. 12060",
		"a2": "B. 12600",
		"a3": "C. 18000",
		"a4": "D. 186000",
		"sl": "C"
	}, {
		"tihao": 26,
		"q": "进程P1、P2、P3、P4和P5的前趋图如下图所示：若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，则需要设置5个信号S1、S2、S3、S4和S5，且信号量S1～S5的初值都等于零。下图中a和b处应分别填（26）；c和d处应分别填写（27）；e和f处应分别填写（28）。",
		"a1": "A. V（S1）P（S2）和V（S3）",
		"a2": "B. P（S1）V（S2）和V（S3）",
		"a3": "C. V（S1）V（S2）和V（S3）",
		"a4": "D. P（S1）P（S2）和V（S3）",
		"sl": "C"
	}, {
		"tihao": 27,
		"q": "进程P1、P2、P3、P4和P5的前趋图如下图所示：若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，则需要设置5个信号S1、S2、S3、S4和S5，且信号量S1～S5的初值都等于零。下图中a和b处应分别填（26）；c和d处应分别填写（27）；e和f处应分别填写（28）。",
		"a1": "A. P（S2）和P（S4）",
		"a2": "B. P（S2）和V（S4）",
		"a3": "C. V（S2）和P（S4）",
		"a4": "D. V（S2）和V（S4）",
		"sl": "B"
	}, {
		"tihao": 28,
		"q": "进程P1、P2、P3、P4和P5的前趋图如下图所示：若用PV操作控制进程P1、P2、P3、P4和P5并发执行的过程，则需要设置5个信号S1、S2、S3、S4和S5，且信号量S1～S5的初值都等于零。下图中a和b处应分别填（26）；c和d处应分别填写（27）；e和f处应分别填写（28）。",
		"a1": "A. P（S4）和V（S4）V（S5）",
		"a2": "B. V（S5）和P（S4）P（S5）",
		"a3": "C. V（S3）和V（S4）V（S5）",
		"a4": "D. P（S3）和P（S4）V（P5）",
		"sl": "B"
	}, {
		"tihao": 29,
		"q": "如下图所示，模块A和模块B都访问相同的全局变量和数据结构，则这两个模块之间的耦合类型为（29）耦合。",
		"a1": "A. 公共",
		"a2": "B. 控制",
		"a3": "C. 标记",
		"a4": "D. 数据",
		"sl": "A"
	}, {
		"tihao": 30,
		"q": "以下关于增量开发模型的叙述中，不正确的是（30）。",
		"a1": "A. 不必等到整个系统开发完成就可以使用",
		"a2": "B. 可以使用较早的增量构件作为原型，从而获得稍后的增量构件需求",
		"a3": "C. 优先级最高的服务先交付，这样最重要的服务接受最多的测试",
		"a4": "D. 有利于进行好的模块划分",
		"sl": "D"
	}, {
		"tihao": 31,
		"q": "在设计软件的模块结构时，（31）不能改进设计质量。",
		"a1": "A. 模块的作用范围应在其控制范围之内",
		"a2": "B. 模块的大小适中",
		"a3": "C. 避免或减少使用病态连接（从中部进入或访问一个模块）",
		"a4": "D. 模块的功能越单纯越好",
		"sl": "D"
	}, {
		"tihao": 32,
		"q": "软件体系结构的各种风格中，仓库风格包含一个数据仓库和若干个其他构件。数据仓库位于该体系结构的中心，其他构件访问该数据仓库并对其中的数据进行增、删、改等操作。以下关于该风格的叙述中，不正确的是（32）。（33）不属于仓库风格。",
		"a1": "A. 支持可更改性和可维护性",
		"a2": "B. 具有可复用的知识源",
		"a3": "C. 支持容错性和健壮性",
		"a4": "D. 测试简单",
		"sl": "D"
	}, {
		"tihao": 33,
		"q": "软件体系结构的各种风格中，仓库风格包含一个数据仓库和若干个其他构件。数据仓库位于该体系结构的中心，其他构件访问该数据仓库并对其中的数据进行增、删、改等操作。以下关于该风格的叙述中，不正确的是（32）。（33）不属于仓库风格。",
		"a1": "A. 数据库系统",
		"a2": "B. 超文本系统",
		"a3": "C. 黑板系统",
		"a4": "D. 编译器",
		"sl": "D"
	}, {
		"tihao": 34,
		"q": "下图（a）所示为一个模块层次结构的例子，图（b）所示为对其进行集成测试的顺序，则此测试采用了（34）测试策略。该测试策略的优点不包括（35）。",
		"a1": "A. 自底向上",
		"a2": "B. 自顶向下",
		"a3": "C. 三明治",
		"a4": "D. 一次性",
		"sl": "C"
	}, {
		"tihao": 35,
		"q": "下图（a）所示为一个模块层次结构的例子，图（b）所示为对其进行集成测试的顺序，则此测试采用了（34）测试策略。该测试策略的优点不包括（35）。",
		"a1": "A. 较早地验证了主要的控制和判断点",
		"a2": "B. 较早地验证了底层模块",
		"a3": "C. 测试的并行程度较高",
		"a4": "D. 较少的驱动模块和桩模块的编写工作量",
		"sl": "D"
	}, {
		"tihao": 36,
		"q": "采用McCabe度量法计算下图所示程序的环路复杂性为（36）。",
		"a1": "A. 1",
		"a2": "B. 2",
		"a3": "C. 3",
		"a4": "D. 4",
		"sl": "C"
	}, {
		"tihao": 37,
		"q": "在面向对象方法中，（37）是父类和子类之间共享数据和方法的机制。子类在原有父类接口的基础上，用适合于自己要求的实现去置换父类中的相应实现称为（38）。",
		"a1": "A. 封装",
		"a2": "B. 继承",
		"a3": "C. 覆盖",
		"a4": "D. 多态",
		"sl": "B"
	}, {
		"tihao": 38,
		"q": "在面向对象方法中，（37）是父类和子类之间共享数据和方法的机制。子类在原有父类接口的基础上，用适合于自己要求的实现去置换父类中的相应实现称为（38）。",
		"a1": "A. 封装",
		"a2": "B. 继承",
		"a3": "C. 覆盖",
		"a4": "D. 多态",
		"sl": "C"
	}, {
		"tihao": 39,
		"q": "在UML用例图中，参与者表示（39）。",
		"a1": "A. 人、硬件或其他系统可以扮演的角色",
		"a2": "B. 可以完成多种动作的相同用户",
		"a3": "C. 不管角色的实际物理用户",
		"a4": "D. 带接口的物理系统或者硬件设计",
		"sl": "A"
	}, {
		"tihao": 40,
		"q": "UML中关联是一个结构关系，描述了一组链。两个类之间（40）关联。",
		"a1": "A. 不能有多个",
		"a2": "B. 可以有多个由不同角色标识的",
		"a3": "C. 可以有任意多个",
		"a4": "D. 的多个关联必须聚合成一个",
		"sl": "B"
	}, {
		"tihao": 41,
		"q": "如下所示的UML图是（41），图中（Ⅰ）表示（42），（Ⅱ）表示（43）。",
		"a1": "A. 序列图",
		"a2": "B. 状态图",
		"a3": "C. 通信图",
		"a4": "D. 活动图",
		"sl": "D"
	}, {
		"tihao": 42,
		"q": "如下所示的UML图是（41），图中（Ⅰ）表示（42），（Ⅱ）表示（43）。",
		"a1": "A. 分叉",
		"a2": "B. 分支",
		"a3": "C. 合并汇合",
		"a4": "D. 流",
		"sl": "A"
	}, {
		"tihao": 43,
		"q": "如下所示的UML图是（41），图中（Ⅰ）表示（42），（Ⅱ）表示（43）。",
		"a1": "A. 分支条件",
		"a2": "B. 监护表达式",
		"a3": "C. 动作名",
		"a4": "D. 流名称",
		"sl": "B"
	}, {
		"tihao": 44,
		"q": "为图形用户界面（GUI）组件定义不同平台的并行类层次结构，适合采用（44）模式。",
		"a1": "A. 享元（Flyweight）",
		"a2": "B. 抽象工厂（Abstract Factory）",
		"a3": "C. 外观（Facade）",
		"a4": "D. 装饰器（Decorator）",
		"sl": "B"
	}, {
		"tihao": 45,
		"q": "（45）设计模式将一个请求封装为一个对象，从而使得可以用不同的请求对客户进行参数化，对请求排队或记录请求日志，以及支持可撤销的操作。",
		"a1": "A. 命令（Command）",
		"a2": "B. 责任链（Chain of Responsibility）",
		"a3": "C. 观察者（Observer）",
		"a4": "D. 策略（Strategy）",
		"sl": "A"
	}, {
		"tihao": 46,
		"q": "（46）设计模式最适合用于发布/订阅消息模型，即当订阅者注册一个主题后，此主题有新消息到来时订阅者就会收到通知。",
		"a1": "A. 适配器（Adapter）",
		"a2": "B. 通知（Notifier）",
		"a3": "C. 观察者（Observer）",
		"a4": "D. 状态（State）",
		"sl": "C"
	}, {
		"tihao": 47,
		"q": "因使用大量的对象而造成很大的存储开销时，适合采用（47）模式进行对象共享，以减少对象数量从而达到较少的内存占用并提升性能。",
		"a1": "A. 组合（Composite）",
		"a2": "B. 享元（Flyweight）",
		"a3": "C. 迭代器（Iterator）",
		"a4": "D. 备忘（Memento）",
		"sl": "B"
	}, {
		"tihao": 48,
		"q": "移进—归约分析法是编译程序（或解释程序）对高级语言源程序进行语法分析的一种方法，属于（48）的语法分析方法。",
		"a1": "A. 自顶向下（或自上而下）",
		"a2": "B. 自底向上（或自下而上）",
		"a3": "C. 自左向右",
		"a4": "D. 自右向左",
		"sl": "B"
	}, {
		"tihao": 49,
		"q": "某确定的有限自动机（DFA）的状态转换图如下图所示（A是初态，C是终态），则该DFA能识别（49）。",
		"a1": "A. aabb",
		"a2": "B. abab",
		"a3": "C. baba",
		"a4": "D. abba",
		"sl": "B"
	}, {
		"tihao": 50,
		"q": "函数main()、f()的定义如下所示，调用函数f()时，第一个参数采用传值（call by value）方式，第二个参数采用传引用（call by reference）方式，main函数中《print(x)》执行后输出的值为（50）。",
		"a1": "A. 1",
		"a2": "B. 6",
		"a3": "C. 11",
		"a4": "D. 12",
		"sl": "D"
	}, {
		"tihao": 51,
		"q": "数据的物理独立性和逻辑独立性分别是通过修改（51）来完成的。",
		"a1": "A. 外模式与内模式之间的映像、模式与内模式之间的映像",
		"a2": "B. 外模式与内模式之间的映像、外模式与模式之间的映像",
		"a3": "C. 外模式与模式之间的映像、模式与内模式之间的映像",
		"a4": "D. 模式与内模式之间的映像、外模式与模式之间的映像",
		"sl": "D"
	}, {
		"tihao": 52,
		"q": "关系规范化在数据库设计的（52）阶段进行。",
		"a1": "A. 需求分析",
		"a2": "B. 概念设计",
		"a3": "C. 逻辑设计",
		"a4": "D. 物理设计",
		"sl": "C"
	}, {
		"tihao": 53,
		"q": "若给定的关系模式为R，U={A,B,C}，F={AB→C,C→B}，则关系R（53）。",
		"a1": "A. 有2个候选关键字AC和BC，并且有3个主属性",
		"a2": "B. 有2个候选关键字AC和AB，并且有3个主属性",
		"a3": "C. 只有一个候选关键字AC，并且有1个非主属性和2个主属性",
		"a4": "D. 只有一个候选关键字AB，并且有1个非主属性和2个主属性",
		"sl": "B"
	}, {
		"tihao": 54,
		"q": "某公司数据库中的元件关系模式为P（元件号，元件名称，供应商，供应商所在地，库存量），函数依赖集F如下所示：F={元件号→元件名称，（元件号，供应商）→库存量，供应商→供应商所在地}元件关系的主键为（54），该关系存在冗余以及插入异常和删除异常等问题。为了解决这一问题需要将元件关系分解（55），分解后的关系模式可以达到（56）。",
		"a1": "A. 元件号，元件名称",
		"a2": "B. 元件号，供应商",
		"a3": "C. 元件号，供应商所在地",
		"a4": "D. 供应商，供应商所在地",
		"sl": "B"
	}, {
		"tihao": 55,
		"q": "某公司数据库中的元件关系模式为P（元件号，元件名称，供应商，供应商所在地，库存量），函数依赖集F如下所示：F={元件号→元件名称，（元件号，供应商）→库存量，供应商→供应商所在地}元件关系的主键为（54），该关系存在冗余以及插入异常和删除异常等问题。为了解决这一问题需要将元件关系分解（55），分解后的关系模式可以达到（56）。",
		"a1": "A. 元件1（元件号，元件名称，库存量）、元件2（供应商，供应商所在地）",
		"a2": "B. 元件1（元件号，元件名称）、元件2（供应商，供应商所在地，库存量）",
		"a3": "C. 元件1（元件号，元件名称）、元件2（元件号，供应商，库存量）、元件3（供应商，供应商所在地）",
		"a4": "D. 元件1（元件号，元件名称）、元件2（元件号，库存量）、元件3（供应商，供应商所在地）、元件4（供应商所在地，库存量）",
		"sl": "C"
	}, {
		"tihao": 56,
		"q": "某公司数据库中的元件关系模式为P（元件号，元件名称，供应商，供应商所在地，库存量），函数依赖集F如下所示：F={元件号→元件名称，（元件号，供应商）→库存量，供应商→供应商所在地}元件关系的主键为（54），该关系存在冗余以及插入异常和删除异常等问题。为了解决这一问题需要将元件关系分解（55），分解后的关系模式可以达到（56）。",
		"a1": "A. 1NF",
		"a2": "B. 2NF",
		"a3": "C. 3NF",
		"a4": "D. 4NF",
		"sl": "C"
	}, {
		"tihao": 57,
		"q": "若元素以a，b，c，d，e的顺序进入一个初始为空的栈中，每个元素进栈、出栈各1次，要求出栈的第一个元素为d，则合法的出栈序列共有（57）种。",
		"a1": "A. 4",
		"a2": "B. 5",
		"a3": "C. 6",
		"a4": "D. 24",
		"sl": "A"
	}, {
		"tihao": 58,
		"q": "设有二叉排序树（或二叉查找树）如下图所示，建立该二叉树的关键码序列不可能是（58）。",
		"a1": "A. 23 31 17 19 11 27 13 90 61",
		"a2": "B. 23 17 19 31 27 90 61 11 13",
		"a3": "C. 23 17 27 19 31 13 11 90 61",
		"a4": "D. 23 31 90 61 27 17 19 11 13",
		"sl": "C"
	}, {
		"tihao": 59,
		"q": "若一棵二叉树的高度（即层数）为h，则该二叉树（59）。",
		"a1": "A. 2h个结点",
		"a2": "B. 有2h-1个结点",
		"a3": "C. 最少有2h-1个结点",
		"a4": "D. 最多有2h-1个结点",
		"sl": "D"
	}, {
		"tihao": 60,
		"q": "在13个元素构成的有序表A[1..13]中进行折半查找（或称为二分查找，向下取整）。那么以下叙述中，错误的是（60）。",
		"a1": "A. 无论要查找哪个元素，都是先与A[7]进行比较",
		"a2": "B. 若要查找的元素等于A[9]，则分别需与A[7]、A[11]、A[9]进行比较",
		"a3": "C. 无论要查找的元素是否在A[]中，最多与表中的4个元素比较即可",
		"a4": "D. 若待查找的元素不在A[]中，最少需要与表中的3个元素进行比较",
		"sl": "B"
	}, {
		"tihao": 61,
		"q": "以下关于图的遍历的叙述中，正确的是（61）。",
		"a1": "A. 图的遍历是从给定的源点出发对每一个顶点仅访问一次的过程",
		"a2": "B. 图的深度优先遍历方法不适用于无向图",
		"a3": "C. 使用队列对图进行广度优先遍历",
		"a4": "D. 图中有回路时则无法进行遍历",
		"sl": "C"
	}, {
		"tihao": 62,
		"q": "考虑一个背包问题，共有n=5个物品，背包容量为W=10，物品的重量和价值分别为：w={2，2，6，5，4}，v={6，3，5，4，6}，求背包问题的最大装包价值。若此为0-1背包问题，分析该问题具有最优子结构，定义递归式为其中c（i，j）表示i个物品、容量为j的0-1背包问题的最大装包价值，最终要求解c（n,W）。采用自底向上的动态规划方法求解，得到最大装包价值为（62），算法的时间复杂度为（63）。若此为部分背包问题，首先采用归并排序算法，根据物品的单位重量价值从大到小排序，然后依次将物品放入背包直至所有物品放入背包中或者背包再无容量，则得到的最大装包价值为（64），算法的时间复杂度为（65）。",
		"a1": "A. 11",
		"a2": "B. 14",
		"a3": "C. 15",
		"a4": "D. 16.67",
		"sl": "C"
	}, {
		"tihao": 63,
		"q": "考虑一个背包问题，共有n=5个物品，背包容量为W=10，物品的重量和价值分别为：w={2，2，6，5，4}，v={6，3，5，4，6}，求背包问题的最大装包价值。若此为0-1背包问题，分析该问题具有最优子结构，定义递归式为其中c（i，j）表示i个物品、容量为j的0-1背包问题的最大装包价值，最终要求解c（n,W）。采用自底向上的动态规划方法求解，得到最大装包价值为（62），算法的时间复杂度为（63）。若此为部分背包问题，首先采用归并排序算法，根据物品的单位重量价值从大到小排序，然后依次将物品放入背包直至所有物品放入背包中或者背包再无容量，则得到的最大装包价值为（64），算法的时间复杂度为（65）。",
		"a1": "A. Θ(nW)",
		"a2": "B. Θ(nlgn)",
		"a3": "C. Θ(n2)",
		"a4": "D. Θ(nlgnW)",
		"sl": "A"
	}, {
		"tihao": 64,
		"q": "考虑一个背包问题，共有n=5个物品，背包容量为W=10，物品的重量和价值分别为：w={2，2，6，5，4}，v={6，3，5，4，6}，求背包问题的最大装包价值。若此为0-1背包问题，分析该问题具有最优子结构，定义递归式为其中c（i，j）表示i个物品、容量为j的0-1背包问题的最大装包价值，最终要求解c（n,W）。采用自底向上的动态规划方法求解，得到最大装包价值为（62），算法的时间复杂度为（63）。若此为部分背包问题，首先采用归并排序算法，根据物品的单位重量价值从大到小排序，然后依次将物品放入背包直至所有物品放入背包中或者背包再无容量，则得到的最大装包价值为（64），算法的时间复杂度为（65）。",
		"a1": "A. 11",
		"a2": "B. 14",
		"a3": "C. 15",
		"a4": "D. 16.67",
		"sl": "D"
	}, {
		"tihao": 65,
		"q": "考虑一个背包问题，共有n=5个物品，背包容量为W=10，物品的重量和价值分别为：w={2，2，6，5，4}，v={6，3，5，4，6}，求背包问题的最大装包价值。若此为0-1背包问题，分析该问题具有最优子结构，定义递归式为其中c（i，j）表示i个物品、容量为j的0-1背包问题的最大装包价值，最终要求解c（n,W）。采用自底向上的动态规划方法求解，得到最大装包价值为（62），算法的时间复杂度为（63）。若此为部分背包问题，首先采用归并排序算法，根据物品的单位重量价值从大到小排序，然后依次将物品放入背包直至所有物品放入背包中或者背包再无容量，则得到的最大装包价值为（64），算法的时间复杂度为（65）。",
		"a1": "A. Θ(nW)",
		"a2": "B. Θ(nlgn)",
		"a3": "C. Θ(n2)",
		"a4": "D. Θ(nlgnW)",
		"sl": "B"
	}, {
		"tihao": 66,
		"q": "默认情况下，FTP服务器的控制端口为（66），上传文件时的端口为（67）。",
		"a1": "A. 大于1024的端口",
		"a2": "B. 20",
		"a3": "C. 80",
		"a4": "D. 21",
		"sl": "D"
	}, {
		"tihao": 67,
		"q": "默认情况下，FTP服务器的控制端口为（66），上传文件时的端口为（67）。",
		"a1": "A. 大于1024的端口",
		"a2": "B. 20",
		"a3": "C. 80",
		"a4": "D. 21",
		"sl": "B"
	}, {
		"tihao": 68,
		"q": "使用ping命令可以进行网络检测，在进行一系列检测时，按照由近及远原则,首先执行的是（68）。",
		"a1": "A. ping默认网关",
		"a2": "B. ping本地IP",
		"a3": "C. ping127.0.0.1",
		"a4": "D. ping远程主机",
		"sl": "C"
	}, {
		"tihao": 69,
		"q": "某PC的Internet协议属性参数如下图所示，默认网关的IP地址是（69）。 ",
		"a1": "A. 8.8.8.8",
		"a2": "B. 202.117.115.3",
		"a3": "C. 192.168.2.254",
		"a4": "D. 202.117.115.18",
		"sl": "C"
	}, {
		"tihao": 70,
		"q": "在下图的SNMP配置中，能够响应Manager2的getRequest请求的是（70）。",
		"a1": "A. Agent1",
		"a2": "B. Agent2",
		"a3": "C. Agent3",
		"a4": "D. Agent4",
		"sl": "A"
	}, {
		"tihao": 71,
		"q": "In the fields of physical security and information security, access control is the selective restriction of access to a place or other resource. The act of accessing may mean consuming, entering, or using. Permission to access a resource is called authorization （授权）．An access control mechanism （71） between a user (or a process executing on behalf of a user) and system resources, such as applications, operating systems, firewalls, routers, files, and databases. The system must first authenticate（验证）a user seeking access. Typically the authentication function determines whether the user is （72） to access the system at all. Then the access control function determines if the specific requested access by this user is permitted. A security administrator maintains an authorization database that specifies what type of access to which resources is allowed for this user. The access control function consults this database to determine whether to（73）access. An auditing function monitors and keeps a record of user accesses to system resources.In practice, a number of（74）may cooperatively share the access control function. All operating systems have at least a rudimentary（基本的）, and in many cases a quite robust, access control component. Add-on security packages can add to the（75）access control capabilities of the OS. Particular applications or utilities, such as a database management system, also incorporate access control functions. External devices, such as firewalls, can also provide access control services.",
		"a1": "A. cooperates",
		"a2": "B. coordinates",
		"a3": "C. connects",
		"a4": "D. mediates",
		"sl": "D"
	}, {
		"tihao": 72,
		"q": "In the fields of physical security and information security, access control is the selective restriction of access to a place or other resource. The act of accessing may mean consuming, entering, or using. Permission to access a resource is called authorization （授权）．An access control mechanism （71） between a user (or a process executing on behalf of a user) and system resources, such as applications, operating systems, firewalls, routers, files, and databases. The system must first authenticate（验证）a user seeking access. Typically the authentication function determines whether the user is （72） to access the system at all. Then the access control function determines if the specific requested access by this user is permitted. A security administrator maintains an authorization database that specifies what type of access to which resources is allowed for this user. The access control function consults this database to determine whether to（73）access. An auditing function monitors and keeps a record of user accesses to system resources.In practice, a number of（74）may cooperatively share the access control function. All operating systems have at least a rudimentary（基本的）, and in many cases a quite robust, access control component. Add-on security packages can add to the（75）access control capabilities of the OS. Particular applications or utilities, such as a database management system, also incorporate access control functions. External devices, such as firewalls, can also provide access control services.",
		"a1": "A. denied",
		"a2": "B. permitted",
		"a3": "C. prohibited",
		"a4": "D. rejected",
		"sl": "B"
	}, {
		"tihao": 73,
		"q": "In the fields of physical security and information security, access control is the selective restriction of access to a place or other resource. The act of accessing may mean consuming, entering, or using. Permission to access a resource is called authorization （授权）．An access control mechanism （71） between a user (or a process executing on behalf of a user) and system resources, such as applications, operating systems, firewalls, routers, files, and databases. The system must first authenticate（验证）a user seeking access. Typically the authentication function determines whether the user is （72） to access the system at all. Then the access control function determines if the specific requested access by this user is permitted. A security administrator maintains an authorization database that specifies what type of access to which resources is allowed for this user. The access control function consults this database to determine whether to（73）access. An auditing function monitors and keeps a record of user accesses to system resources.In practice, a number of（74）may cooperatively share the access control function. All operating systems have at least a rudimentary（基本的）, and in many cases a quite robust, access control component. Add-on security packages can add to the（75）access control capabilities of the OS. Particular applications or utilities, such as a database management system, also incorporate access control functions. External devices, such as firewalls, can also provide access control services.",
		"a1": "A. open",
		"a2": "B. monitor",
		"a3": "C. grant",
		"a4": "D. seek",
		"sl": "C"
	}, {
		"tihao": 74,
		"q": "In the fields of physical security and information security, access control is the selective restriction of access to a place or other resource. The act of accessing may mean consuming, entering, or using. Permission to access a resource is called authorization （授权）．An access control mechanism （71） between a user (or a process executing on behalf of a user) and system resources, such as applications, operating systems, firewalls, routers, files, and databases. The system must first authenticate（验证）a user seeking access. Typically the authentication function determines whether the user is （72） to access the system at all. Then the access control function determines if the specific requested access by this user is permitted. A security administrator maintains an authorization database that specifies what type of access to which resources is allowed for this user. The access control function consults this database to determine whether to（73）access. An auditing function monitors and keeps a record of user accesses to system resources.In practice, a number of（74）may cooperatively share the access control function. All operating systems have at least a rudimentary（基本的）, and in many cases a quite robust, access control component. Add-on security packages can add to the（75）access control capabilities of the OS. Particular applications or utilities, such as a database management system, also incorporate access control functions. External devices, such as firewalls, can also provide access control services.",
		"a1": "A. components",
		"a2": "B. users",
		"a3": "C. mechanisms",
		"a4": "D. algorithms",
		"sl": "A"
	}, {
		"tihao": 75,
		"q": "In the fields of physical security and information security, access control is the selective restriction of access to a place or other resource. The act of accessing may mean consuming, entering, or using. Permission to access a resource is called authorization （授权）．An access control mechanism （71） between a user (or a process executing on behalf of a user) and system resources, such as applications, operating systems, firewalls, routers, files, and databases. The system must first authenticate（验证）a user seeking access. Typically the authentication function determines whether the user is （72） to access the system at all. Then the access control function determines if the specific requested access by this user is permitted. A security administrator maintains an authorization database that specifies what type of access to which resources is allowed for this user. The access control function consults this database to determine whether to（73）access. An auditing function monitors and keeps a record of user accesses to system resources.In practice, a number of（74）may cooperatively share the access control function. All operating systems have at least a rudimentary（基本的）, and in many cases a quite robust, access control component. Add-on security packages can add to the（75）access control capabilities of the OS. Particular applications or utilities, such as a database management system, also incorporate access control functions. External devices, such as firewalls, can also provide access control services.",
		"a1": "A. remote",
		"a2": "B. native",
		"a3": "C. controlled",
		"a4": "D. automated",
		"sl": "B"
	}]
}
软件设计16年下={
	"code": 0,
	"list": [{
		"tihao": 1,
		"q": "在程序运行过程中，CPU需要将指令从内存中取出并加以分析和执行。CPU依据（1）来区分在内存中以二进制编码形式存放的指令和数据。",
		"a1": "A. 指令周期的不同阶段",
		"a2": "B. 指令和数据的寻址方式",
		"a3": "C. 指令操作码的译码结果",
		"a4": "D. 指令和数据所在的存储单元",
		"sl": "A"
	}, {
		"tihao": 2,
		"q": "计算机在一个指令周期的过程中，为从内存读取指令操作码，首先要将（2）的内容送到地址总线上。",
		"a1": "A. 指令寄存器（IR）",
		"a2": "B. 通用寄存器（GR）",
		"a3": "C. 程序计数器（PC）",
		"a4": "D. 状态寄存器（PSW）",
		"sl": "C"
	}, {
		"tihao": 3,
		"q": "设16位浮点数，其中阶符1位、阶码值6位、数符1位、尾数8位。若阶码用移码表示，尾数用补码表示，则该浮点数所能表示的数值范围是（3）。",
		"a1": "A. -264～（1-2-8）264",
		"a2": "B. -263～（1-2-8）263",
		"a3": "C. -264～（1-2-（1-2-8）264～（1-2-8）264",
		"a4": "D. -（1-2-8）263～（1-2-8）263",
		"sl": "B"
	}, {
		"tihao": 4,
		"q": "已知数据信息为16位，最少应附加（4）位校验位，以实现海明码纠错。",
		"a1": "A. 3",
		"a2": "B. 4",
		"a3": "C. 5",
		"a4": "D. 6",
		"sl": "C"
	}, {
		"tihao": 5,
		"q": "将一条指令的执行过程分解为取址、分析和执行三步，按照流水方式执行，若取址时间t取址=4△t、分析时间t分析=2△t、执行时间t执行=3△t，则执行完100条指令，需要的时间为（5）△t。",
		"a1": "A. 200",
		"a2": "B. 300",
		"a3": "C. 400",
		"a4": "D. 405",
		"sl": "D"
	}, {
		"tihao": 6,
		"q": "以下关于Cache与主存间地址映射的叙述中，正确的是（6）。",
		"a1": "A. 操作系统负责管理Cache与主存之间的地址映射",
		"a2": "B. 程序员需要通过编程来处理Cache与主存之间的地址映射",
		"a3": "C. 应用软件对Cache与主存之间的地址映射进行调度",
		"a4": "D. 由硬件自动完成Cache与主存之间的地址映射",
		"sl": "D"
	}, {
		"tihao": 7,
		"q": "可用于数字签名的算法是（7）。",
		"a1": "A. RSA",
		"a2": "B. IDEA",
		"a3": "C. RC4",
		"a4": "D. MD5",
		"sl": "A"
	}, {
		"tihao": 8,
		"q": "（8）不是数字签名的作用。",
		"a1": "A. 接收者可验证消息来源的真实性",
		"a2": "B. 发送者无法否认发送过该消息",
		"a3": "C. 接收者无法伪造或篡改消息",
		"a4": "D. 可验证接收者合法性",
		"sl": "D"
	}, {
		"tihao": 9,
		"q": "在网络设计和实施过程中要采取多种安全措施，其中（9）是针对系统安全需求的措施。",
		"a1": "A. 设备防雷击",
		"a2": "B. 入侵检测",
		"a3": "C. 漏洞发现与补丁管理",
		"a4": "D. 流量控制",
		"sl": "C"
	}, {
		"tihao": 10,
		"q": "（10）的保护期限是可以延长的。",
		"a1": "A. 专利权",
		"a2": "B. 商标权",
		"a3": "C. 著作权",
		"a4": "D. 商业秘密权",
		"sl": "B"
	}, {
		"tihao": 11,
		"q": "甲公司软件设计师完成了一项涉及计算机程序的发明。之后，乙公司软件设计师也完成了与甲公司软件设计师相同的涉及计算机程序的发明。甲、乙公司于同一天向专利局申请发明专利。此情形下，（11）是专利权申请人。",
		"a1": "A. 甲公司",
		"a2": "B. 甲、乙两公司",
		"a3": "C. 乙公司",
		"a4": "D. 由甲、乙公司协商确定的公司",
		"sl": "D"
	}, {
		"tihao": 13,
		"q": "在FM方式的数字音乐合成器中，改变数字载波频率可以改变乐音的（13），改变它的信号幅度可以改变乐音的（14）。",
		"a1": "A. 音调",
		"a2": "B. 音色",
		"a3": "C. 音高",
		"a4": "D. 音质",
		"sl": "A"
	}, {
		"tihao": 14,
		"q": "在FM方式的数字音乐合成器中，改变数字载波频率可以改变乐音的（13），改变它的信号幅度可以改变乐音的（14）。",
		"a1": "A. 音调",
		"a2": "B. 音域",
		"a3": "C. 音高",
		"a4": "D. 带宽",
		"sl": "C"
	}, {
		"tihao": 15,
		"q": "结构化开发方法中，（15）主要包含对数据结构和算法的设计。",
		"a1": "A. 体系结构设计",
		"a2": "B. 数据设计",
		"a3": "C. 接口设计",
		"a4": "D. 过程设计",
		"sl": "D"
	}, {
		"tihao": 16,
		"q": "在敏捷过程的开发方法中，（16）使用了迭代的方法，其中，把每段时间（30天）一次的迭代称为一个《冲刺》，并按需求的优先级别来实现产品，多个自组织和自治的小组并行地递增实现产品。",
		"a1": "A. 极限编程XP",
		"a2": "B. 水晶法",
		"a3": "C. 并列争球法",
		"a4": "D. 自适应软件开发",
		"sl": "C"
	}, {
		"tihao": 17,
		"q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示相应活动的持续时间（天），则完成该项目的最少时间为（17）天。活动BC和BF最多可以晚开始（18）天而不会影响整个项目的进度。",
		"a1": "A. 11",
		"a2": "B. 15",
		"a3": "C. 16",
		"a4": "D. 18",
		"sl": "D"
	}, {
		"tihao": 18,
		"q": "某软件项目的活动图如下图所示，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，边上的数字表示相应活动的持续时间（天），则完成该项目的最少时间为（17）天。活动BC和BF最多可以晚开始（18）天而不会影响整个项目的进度。",
		"a1": "A. 0和7",
		"a2": "B. 0和11",
		"a3": "C. 2和7",
		"a4": "D. 2和11",
		"sl": "A"
	}, {
		"tihao": 19,
		"q": "成本估算时，（19）方法以规模作为成本的主要因素，考虑多个成本驱动因子。该方法包括三个阶段性模型，即应用组装模型、早期设计阶段模型和体系结构阶段模型。",
		"a1": "A. 专家估算",
		"a2": "B. Wolverton",
		"a3": "C. COCOMO",
		"a4": "D. COCOMO Ⅱ",
		"sl": "D"
	}, {
		"tihao": 21,
		"q": "常用的函数参数传递方式有传值与传引用两种。（21）。",
		"a1": "A. 在传值方式下，形参与实参之间互相传值",
		"a2": "B. 在传值方式下，实参不能是变量",
		"a3": "C. 在传引用方式下，修改形参实质上改变了实参的值。",
		"a4": "D. 在传引用方式下，实参可以是任意的变量和表达式。",
		"sl": "C"
	}, {
		"tihao": 22,
		"q": "二维数组a[1..N，1..N]可以按行存储或按列存储。对于数组元素a[i,j]（1<=i,j<=N），当（22）时，在按行和按列两种存储方式下，其偏移量相同。",
		"a1": "A. i≠j",
		"a2": "B. i=j",
		"a3": "C. i>j",
		"a4": "D. i<j",
		"sl": "B"
	}, {
		"tihao": 23,
		"q": "实时操作系统主要用于有实时要求的过程控制等领域。实时系统对于来自外部的事件必须在（23）。",
		"a1": "A. 一个时间片内进行处理",
		"a2": "B. 一个周转时间内进行处理",
		"a3": "C. 一个机器周期内进行处理",
		"a4": "D. 被控对象规定的时间内做出及时响应并对其进行处理",
		"sl": "D"
	}, {
		"tihao": 24,
		"q": "假设某计算机系统中只有一个CPU、一台输入设备和一台输出设备，若系统中有四个作业T1、T2、T3和T4，系统采用优先级调度，且T1的优先级>T2的优先级>T3的优先级>T4的优先级。每个作业Ti具有三个程序段：输入Ii、计算Ci和输出Pi（i=1，2，3，4），其执行顺序为Ii→Ci→Pi。这四个作业各程序段并发执行的前驱图如下所示。图中①、②分别为（24），③、④、⑤分别为（25）。",
		"a1": "A. l2、P2",
		"a2": "B. I2、C2",
		"a3": "C. C1、P2",
		"a4": "D. C1、P3",
		"sl": "C"
	}, {
		"tihao": 25,
		"q": "假设某计算机系统中只有一个CPU、一台输入设备和一台输出设备，若系统中有四个作业T1、T2、T3和T4，系统采用优先级调度，且T1的优先级>T2的优先级>T3的优先级>T4的优先级。每个作业Ti具有三个程序段：输入Ii、计算Ci和输出Pi（i=1，2，3，4），其执行顺序为Ii→Ci→Pi。这四个作业各程序段并发执行的前驱图如下所示。图中①、②分别为（24），③、④、⑤分别为（25）。",
		"a1": "A. C2、C4、P4",
		"a2": "B. l2、l3、C4",
		"a3": "C. I3、P3、P4",
		"a4": "D. l3、C4、P4",
		"sl": "D"
	}, {
		"tihao": 26,
		"q": "假设段页式存储管理系统中的地址结构如下图所示，则系统（26）。",
		"a1": "A. 最多可有256个段，每个段的大小均为2048个页，页的大小为8K",
		"a2": "B. 最多可有256个段，每个段最大允许有2048个页，页的大小为8K",
		"a3": "C. 最多可有512个段，每个段的大小均为1024个页，页的大小为4K",
		"a4": "D. 最多可有512个段，每个段最大允许有1024个页，页的大小为4K",
		"sl": "B"
	}, {
		"tihao": 27,
		"q": "假设系统中有n个进程共享3台扫描仪，并采用PV操怍实现进程同步与互斥。若系统信号量S的当前值为-1，进程P1、P2又分别执行了1次P（S）操作，那么信号量S的值应为（27）。",
		"a1": "A. 3",
		"a2": "B. -3",
		"a3": "C. 1",
		"a4": "D. -1",
		"sl": "B"
	}, {
		"tihao": 28,
		"q": "某字长为32位的计算机的文件管理系统采用位示图（bitmap）记录磁盘的使用情况。若磁盘的容量为300GB，物理块的大小为1MB，那么位示图的大小为（28）个字。",
		"a1": "A. 1200",
		"a2": "B. 3200",
		"a3": "C. 6400",
		"a4": "D. 9600",
		"sl": "D"
	}, {
		"tihao": 29,
		"q": "某开发小组欲为一公司开发一个产品控制软件，监控产品的生产和销售过程，从购买各种材料开始，到产品的加工和销售进行全程跟踪。购买材料的流程、产品的加工过程以及销售过程可能会发生变化。该软件的开发最不适宜采用（29）模型，主要是因为这种模型（30）。",
		"a1": "A. 瀑布",
		"a2": "B. 原型",
		"a3": "C. 增量",
		"a4": "D. 喷泉",
		"sl": "A"
	}, {
		"tihao": 30,
		"q": "某开发小组欲为一公司开发一个产品控制软件，监控产品的生产和销售过程，从购买各种材料开始，到产品的加工和销售进行全程跟踪。购买材料的流程、产品的加工过程以及销售过程可能会发生变化。该软件的开发最不适宜采用（29）模型，主要是因为这种模型（30）。",
		"a1": "A. 不能解决风险",
		"a2": "B. 不能快速提交软件",
		"a3": "C. 难以适应变化的需求",
		"a4": "D. 不能理解用户的需求",
		"sl": "C"
	}, {
		"tihao": 31,
		"q": "（31）不属于软件质量特性中的可移植性。",
		"a1": "A. 适应性",
		"a2": "B. 易安装性",
		"a3": "C. 易替换性",
		"a4": "D. 易理解性",
		"sl": "D"
	}, {
		"tihao": 32,
		"q": "对下图所示流程图采用白盒测试方法进行测试，若要满足路径覆盖，则至少需要（32）个测试用例。采用McCabe度量法计算该程序的环路复杂性为（33）。",
		"a1": "A. 3",
		"a2": "B. 4",
		"a3": "C. 6",
		"a4": "D. 8",
		"sl": "C"
	}, {
		"tihao": 33,
		"q": "对下图所示流程图采用白盒测试方法进行测试，若要满足路径覆盖，则至少需要（32）个测试用例。采用McCabe度量法计算该程序的环路复杂性为（33）。",
		"a1": "A. 1",
		"a2": "B. 2",
		"a3": "C. 3",
		"a4": "D. 4",
		"sl": "D"
	}, {
		"tihao": 34,
		"q": "计算机系统的（34）可以用MTBF/（1+MTBF）来度量，其中MTBF为平均失效间隔时间。",
		"a1": "A. 可靠性",
		"a2": "B. 可用性",
		"a3": "C. 可维护性",
		"a4": "D. 健壮性",
		"sl": "B"
	}, {
		"tihao": 35,
		"q": "以下关于软件测试的叙述中，不正确的是（35）。",
		"a1": "A. 在设计测试用例时应考虑输入数据和预期输出结果",
		"a2": "B. 软件测试的目的是证明软件的正确性",
		"a3": "C. 在设计测试用例时，应该包括合理的输入条件",
		"a4": "D. 在设计测试用例时，应该包括不合理的输入条件",
		"sl": "B"
	}, {
		"tihao": 36,
		"q": "某模块中有两个处理A和B，分别对数据结构X写数据和读数据，则该模块的内聚类型为（36）内聚。",
		"a1": "A. 逻辑",
		"a2": "B. 过程",
		"a3": "C. 通信",
		"a4": "D. 内容",
		"sl": "C"
	}, {
		"tihao": 37,
		"q": "在面向对象方法中，不同对象收到同一消息可以产生完全不同的结果，这一现象称为（37）。在使用时，用户可以发送一个通用的消息，而实现的细节则由接收对象自行决定。",
		"a1": "A. 接口",
		"a2": "B. 继承",
		"a3": "C. 覆盖",
		"a4": "D. 多态",
		"sl": "D"
	}, {
		"tihao": 38,
		"q": "在面向对象方法中，支持多态的是（38）。",
		"a1": "A. 静态分配",
		"a2": "B. 动态分配",
		"a3": "C. 静态类型",
		"a4": "D. 动态绑定",
		"sl": "D"
	}, {
		"tihao": 39,
		"q": "面向对象分析的目的是为了获得对应用问题的理解，其主要活动不包括（39）。",
		"a1": "A. 认定并组织对象",
		"a2": "B. 描述对象间的相互作用",
		"a3": "C. 面向对象程序设计",
		"a4": "D. 确定基于对象的操作",
		"sl": "C"
	}, {
		"tihao": 40,
		"q": "如下所示的UML状态图中，（40）时，不一定会离开状态B。",
		"a1": "A. 状态B中的两个结束状态均达到",
		"a2": "B. 在当前状态为B2时，事件e2发生",
		"a3": "C. 事件e2发生",
		"a4": "D. 事件e1发生",
		"sl": "C"
	}, {
		"tihao": 41,
		"q": "以下关于UML状态图中转换（transition）的叙述中，不正确的是（41）。",
		"a1": "A. 活动可以在转换时执行也可以在状态内执行",
		"a2": "B. 监护条件只有在相应的事件发生时才进行检查",
		"a3": "C. 一个转换可以有事件触发器、监护条件和一个状态",
		"a4": "D. 事件触发转换",
		"sl": "C"
	}, {
		"tihao": 42,
		"q": "下图①②③④所示是UML（42）。现有场景：一名医生（Doctor）可以治疗多位病人（Patient），一位病人可以由多名医生治疗，一名医生可能多次治疗同一位病人。要记录哪名医生治疗哪位病人时，需要存储治疗（Treatment）的日期和时间。以下①②③④图中（43）。是描述此场景的模型。",
		"a1": "A. 用例图",
		"a2": "B. 对象图",
		"a3": "C. 类图",
		"a4": "D. 协作图",
		"sl": "C"
	}, {
		"tihao": 43,
		"q": "下图①②③④所示是UML（42）。现有场景：一名医生（Doctor）可以治疗多位病人（Patient），一位病人可以由多名医生治疗，一名医生可能多次治疗同一位病人。要记录哪名医生治疗哪位病人时，需要存储治疗（Treatment）的日期和时间。以下①②③④图中（43）。是描述此场景的模型。",
		"a1": "A. ①",
		"a2": "B. ②",
		"a3": "C. ③",
		"a4": "D. ④",
		"sl": "A"
	}, {
		"tihao": 44,
		"q": "（44）模式定义一系列的算法，把它们一个个封装起来，并且使它们可以相互替换，使得算法可以独立于使用它们的客户而变化。以下（45）情况适合选用该模式。①一个客户需要使用一组相关对象②一个对象的改变需要改变其它对象 ③需要使用一个算法的不同变体④许多相关的类仅仅是行为有异",
		"a1": "A. 命令（Command）",
		"a2": "B. 责任链（Chain of Responsibility）",
		"a3": "C. 观察者（Observer）",
		"a4": "D. 策略（Strategy）",
		"sl": "D"
	}, {
		"tihao": 45,
		"q": "（44）模式定义一系列的算法，把它们一个个封装起来，并且使它们可以相互替换，使得算法可以独立于使用它们的客户而变化。以下（45）情况适合选用该模式。①一个客户需要使用一组相关对象②一个对象的改变需要改变其它对象 ③需要使用一个算法的不同变体④许多相关的类仅仅是行为有异",
		"a1": "A. ①②",
		"a2": "B. ②③",
		"a3": "C. ③④",
		"a4": "D. ①④",
		"sl": "C"
	}, {
		"tihao": 46,
		"q": "（46）模式将一个复杂对象的构建与其表示分离，使得同样的构建过程可以创建不同的表示。以下（47）情况适合选用该模式。①抽象复杂对象的构建步骤②基于构建过程的具体实现构建复杂对象的不同表示③一个类仅有一个实例④一个类的实例只能有几个不同状态组合中的一种",
		"a1": "A. 生成器（Builder）",
		"a2": "B. 工厂方法（Factory Method）",
		"a3": "C. 原型（Prototype）",
		"a4": "D. 单例（Singleton）",
		"sl": "A"
	}, {
		"tihao": 47,
		"q": "（46）模式将一个复杂对象的构建与其表示分离，使得同样的构建过程可以创建不同的表示。以下（47）情况适合选用该模式。①抽象复杂对象的构建步骤②基于构建过程的具体实现构建复杂对象的不同表示③一个类仅有一个实例④一个类的实例只能有几个不同状态组合中的一种",
		"a1": "A. ①②",
		"a2": "B. ②③",
		"a3": "C. ③④",
		"a4": "D. ①④",
		"sl": "A"
	}, {
		"tihao": 48,
		"q": "由字符a、b构成的字符串中，若每个a后至少跟一个b，则该字符串集合可用正规式表示为（48）。",
		"a1": "A. （b|ab）*",
		"a2": "B. （ab*）*",
		"a3": "C. （a*b*）*",
		"a4": "D. （a|b）*",
		"sl": "A"
	}, {
		"tihao": 49,
		"q": "乔姆斯基（Chomsky）将文法分为4种类型，程序设计语言的大多数语法现象可用其中的（49）描述。",
		"a1": "A. 上下文有关文法",
		"a2": "B. 上下文无关文法",
		"a3": "C. 正规文法",
		"a4": "D. 短语结构文法",
		"sl": "B"
	}, {
		"tihao": 50,
		"q": "运行下面的C程序代码段，会出现（50）错误。int k=0;for(;k<100;);{k++;}",
		"a1": "A. 变量未定义",
		"a2": "B. 静态语义",
		"a3": "C. 语法",
		"a4": "D. 动态语义",
		"sl": "D"
	}, {
		"tihao": 51,
		"q": "在数据库系统中，一般由DBA使用DBMS提供的授权功能为不同用户授权，其主要目的是为了保证数据库的（51）。",
		"a1": "A. 正确性",
		"a2": "B. 安全性",
		"a3": "C. 一致性",
		"a4": "D. 完整性",
		"sl": "B"
	}, {
		"tihao": 52,
		"q": "给定关系模式R（U,F），其中：U为关系模式R中的属性集，F是U上的一组函数依赖。假设U={A1，A2，A3，A4}，F={A1→A2，A1A2→A3，A1→A4，A2→A4}，那么关系R的主键应为（52）。函数依赖集F中的（53）是冗余的。",
		"a1": "A. A1",
		"a2": "B. A1A2",
		"a3": "C. A1A3",
		"a4": "D. A1A2A3",
		"sl": "B"
	}, {
		"tihao": 53,
		"q": "给定关系模式R（U,F），其中：U为关系模式R中的属性集，F是U上的一组函数依赖。假设U={A1，A2，A3，A4}，F={A1→A2，A1A2→A3，A1→A4，A2→A4}，那么关系R的主键应为（52）。函数依赖集F中的（53）是冗余的。",
		"a1": "A. A1→A2",
		"a2": "B. A1A2→A3",
		"a3": "C. A1→A4",
		"a4": "D. A2→A4",
		"sl": "C"
	}, {
		"tihao": 54,
		"q": "给定关系R（A，B，C，D）和关系S（A，C，E，F），对其进行自然连接运算R⋈S后的属性列为（54）个；与σR.B>S.E(R⋈S)等价的关系代数表达式为（55）。",
		"a1": "A. 4",
		"a2": "B. 5",
		"a3": "C. 6",
		"a4": "D. 8",
		"sl": "C"
	}, {
		"tihao": 55,
		"q": "给定关系R（A，B，C，D）和关系S（A，C，E，F），对其进行自然连接运算R⋈S后的属性列为（54）个；与σR.B>S.E(R⋈S)等价的关系代数表达式为（55）。",
		"a1": "A. σ2<7（R×S）",
		"a2": "B. π1,2,3,4,7,8（σ1=5∧2>7∧3=6（R×S））",
		"a3": "C. σ2<‘7’（R×S）",
		"a4": "D. π1,2,3,4,7,8（σ1=5∧2>‘7’∧3=6（R×S））",
		"sl": "B"
	}, {
		"tihao": 56,
		"q": "下列查询B=《大数据》且F=《开发平台》，结果集属性列为A、B、C、F的关系代数表达式中，查询效率最高的是（56）。",
		"a1": "A. π1,2,3,8（σ2='大数据'^1=5^3=6^8='开发平台'（R×S））",
		"a2": "B. π1,2,3,8（σ1=5^3=6^8='开发平台'（σ2='大数据'（R）×S））",
		"a3": "C. π1,2,3,8（σ2='大数据'^1=5^3=6（R×σ4='开发平台'（S））",
		"a4": "D. π1,2,3,8（σ1=5^3=6（σ2='大数据'（R）×σ4='开发平台'（S）））",
		"sl": "D"
	}, {
		"tihao": 57,
		"q": "拓扑序列是有向无环图中所有顶点的一个线性序列，若有向图中存在弧<v，w>或存在从顶点v到w的路径，则在该有向图的任一拓扑序列中，v一定在w之前。下面有向图的拓扑序列是（57）。",
		"a1": "A. 4 1 2 3 5",
		"a2": "B. 4 3 1 2 5",
		"a3": "C. 4 2 1 3 5",
		"a4": "D. 4 1 3 2 5",
		"sl": "A"
	}, {
		"tihao": 58,
		"q": "设有一个包含n个元素的有序线性表。在等概率情况下删除其中的一个元素，若采用顺序存储结构，则平均需要移动（58）个元素；若采用单链表存储，则平均需要移动（59）个元素。",
		"a1": "A. 1",
		"a2": "B. (n-1)/2",
		"a3": "C. logn",
		"a4": "D. n",
		"sl": "B"
	}, {
		"tihao": 59,
		"q": "设有一个包含n个元素的有序线性表。在等概率情况下删除其中的一个元素，若采用顺序存储结构，则平均需要移动（58）个元素；若采用单链表存储，则平均需要移动（59）个元素。",
		"a1": "A. 0",
		"a2": "B. 1",
		"a3": "C. (n-1)/2",
		"a4": "D. n/2",
		"sl": "A"
	}, {
		"tihao": 60,
		"q": "具有3个结点的二叉树有（60）种形态。",
		"a1": "A. 2",
		"a2": "B. 3",
		"a3": "C. 5",
		"a4": "D. 7",
		"sl": "C"
	}, {
		"tihao": 61,
		"q": "以下关于二叉排序树（或二叉查找树、二叉搜索树）的叙述中，正确的是（61） 。",
		"a1": "A. 对二叉排序树进行先序、中序和后序遍历，都得到结点关键字的有序序列",
		"a2": "B. 含有n个结点的二叉排序树高度为（log2n）+1",
		"a3": "C. 从根到任意一个叶子结点的路径上，结点的关键字呈现有序排列的特点",
		"a4": "D. 从左到右排列同层次的结点，其关键字呈现有序排列的特点",
		"sl": "D"
	}, {
		"tihao": 62,
		"q": "下表为某文件中字符的出现频率，采用霍夫曼编码对下列字符编码，则字符序列《bee》的编码为（62）；编码《110001001101》的对应的字符序列为（63）。",
		"a1": "A. 10111011101",
		"a2": "B. 10111001100",
		"a3": "C. 001100100",
		"a4": "D. 110011011",
		"sl": "A"
	}, {
		"tihao": 63,
		"q": "下表为某文件中字符的出现频率，采用霍夫曼编码对下列字符编码，则字符序列《bee》的编码为（62）；编码《110001001101》的对应的字符序列为（63）。",
		"a1": "A. bad",
		"a2": "B. bee",
		"a3": "C. face",
		"a4": "D. bace",
		"sl": "C"
	}, {
		"tihao": 64,
		"q": "两个矩阵Am*n和Bn*p相乘，用基本的方法进行，则需要的乘法次数为m*n*p。多个矩阵相乘满足结合律，不同的乘法顺序所需要的乘法次数不同。考虑采用动态规划方法确定Mi，M(i+1)，…，Mj多个矩阵连乘的最优顺序，即所需要的乘法次数最少。最少乘法次数用m[i,j]表示，其递归式定义为：其中i、j和k为矩阵下标，矩阵序列中Mi的维度为（pi-1）*pi采用自底向上的方法实现该算法来确定n个矩阵相乘的顺序，其时间复杂度为（64）。若四个矩阵M1、 M2、M3、M4相乘的维度序列为2、6、3、10、3，采用上述算法求解，则乘法次数为（65）。",
		"a1": "A. O（n2）",
		"a2": "B. O（n2lgn）",
		"a3": "C. O（n3）",
		"a4": "D. O（n3lgn）",
		"sl": "C"
	}, {
		"tihao": 65,
		"q": "两个矩阵Am*n和Bn*p相乘，用基本的方法进行，则需要的乘法次数为m*n*p。多个矩阵相乘满足结合律，不同的乘法顺序所需要的乘法次数不同。考虑采用动态规划方法确定Mi，M(i+1)，…，Mj多个矩阵连乘的最优顺序，即所需要的乘法次数最少。最少乘法次数用m[i,j]表示，其递归式定义为：其中i、j和k为矩阵下标，矩阵序列中Mi的维度为（pi-1）*pi采用自底向上的方法实现该算法来确定n个矩阵相乘的顺序，其时间复杂度为（64）。若四个矩阵M1、 M2、M3、M4相乘的维度序列为2、6、3、10、3，采用上述算法求解，则乘法次数为（65）。",
		"a1": "A. 156",
		"a2": "B. 144",
		"a3": "C. 180",
		"a4": "D. 360",
		"sl": "B"
	}, {
		"tihao": 66,
		"q": "以下协议中属于应用层协议的是（66），该协议的报文封装在（67）。",
		"a1": "A. SNMP",
		"a2": "B. ARP",
		"a3": "C. ICMP",
		"a4": "D. X.25",
		"sl": "A"
	}, {
		"tihao": 67,
		"q": "以下协议中属于应用层协议的是（66），该协议的报文封装在（67）。",
		"a1": "A. TCP",
		"a2": "B. IP",
		"a3": "C. UDP",
		"a4": "D. ICMP",
		"sl": "C"
	}, {
		"tihao": 68,
		"q": "某公司内部使用wb.xyz.com.cn作为访问某服务器的地址，其中wb是（67）。",
		"a1": "A. 主机名",
		"a2": "B. 协议名",
		"a3": "C. 目录名",
		"a4": "D. 文件名",
		"sl": "A"
	}, {
		"tihao": 69,
		"q": "如果路由器收到了多个路由协议转发的关于某个目标的多条路由，那么决定采用哪条路由的策略是（69）。",
		"a1": "A. 选择与自己路由协议相同的",
		"a2": "B. 选择路由费用最小的",
		"a3": "C. 比较各个路由的管理距离",
		"a4": "D. 比较各个路由协议的版本",
		"sl": "C"
	}, {
		"tihao": 70,
		"q": "与地址220.112.179.92匹配的路由表的表项是（70）。",
		"a1": "A. 220.112.145.32/22",
		"a2": "B. 220.112.145.64/22",
		"a3": "C. 220.112.147.64/22",
		"a4": "D. 220.112.177.64/22",
		"sl": "D"
	}, {
		"tihao": 71,
		"q": "Software entities are more complex for their size than perhaps any other human construct, because no two parts are alike (at least above the statement level). If they are, we make the two similar parts into one, a（71）, open or closed. In this respect software systems differ profoundly from computers,buildings, or automobiles, where repeated elements abound.Digital computers are themselves more complex than most things people build; they have very large numbers of states. This makes conceiving, describing, and testing them hard. Software systems have orders of magnitude more （72）than computers do.Likewise, a scaling-up of a software entity is not merely a repetition of the same elements in larger size; it is necessarily an increase in the number of different elements. In most cases, the elements interact with each other in some（73）fashion,and the complexity of the whole increases much more than linearly.The complexity of software is a(an)（74）property, not an accidental one. Hence descriptions of a software entity that abstract away its complexity often abstract away its essence.Mathematics and the physical sciences made great strides for three centuries by constructing simplified models of complex phenomena, deriving properties from the models, and verifying those properties experimentally. This worked because the complexities（75）in the models were not the essential properties of the phenomena. It does not work when the complexities are the essence.Many of the classical problems of developing software products derive from this essential complexity and its nonlinear increases with size. Not only technical problems but management problems as well come from the complexity.",
		"a1": "A. task",
		"a2": "B. job",
		"a3": "C. subroutine",
		"a4": "D. program",
		"sl": "C"
	}, {
		"tihao": 72,
		"q": "Software entities are more complex for their size than perhaps any other human construct, because no two parts are alike (at least above the statement level). If they are, we make the two similar parts into one, a（71）, open or closed. In this respect software systems differ profoundly from computers,buildings, or automobiles, where repeated elements abound.Digital computers are themselves more complex than most things people build; they have very large numbers of states. This makes conceiving, describing, and testing them hard. Software systems have orders of magnitude more （72）than computers do.Likewise, a scaling-up of a software entity is not merely a repetition of the same elements in larger size; it is necessarily an increase in the number of different elements. In most cases, the elements interact with each other in some（73）fashion,and the complexity of the whole increases much more than linearly.The complexity of software is a(an)（74）property, not an accidental one. Hence descriptions of a software entity that abstract away its complexity often abstract away its essence.Mathematics and the physical sciences made great strides for three centuries by constructing simplified models of complex phenomena, deriving properties from the models, and verifying those properties experimentally. This worked because the complexities（75）in the models were not the essential properties of the phenomena. It does not work when the complexities are the essence.Many of the classical problems of developing software products derive from this essential complexity and its nonlinear increases with size. Not only technical problems but management problems as well come from the complexity.",
		"a1": "A. states",
		"a2": "B. parts",
		"a3": "C. conditions",
		"a4": "D. expressions",
		"sl": "A"
	}, {
		"tihao": 73,
		"q": "Software entities are more complex for their size than perhaps any other human construct, because no two parts are alike (at least above the statement level). If they are, we make the two similar parts into one, a（71）, open or closed. In this respect software systems differ profoundly from computers,buildings, or automobiles, where repeated elements abound.Digital computers are themselves more complex than most things people build; they have very large numbers of states. This makes conceiving, describing, and testing them hard. Software systems have orders of magnitude more （72）than computers do.Likewise, a scaling-up of a software entity is not merely a repetition of the same elements in larger size; it is necessarily an increase in the number of different elements. In most cases, the elements interact with each other in some（73）fashion,and the complexity of the whole increases much more than linearly.The complexity of software is a(an)（74）property, not an accidental one. Hence descriptions of a software entity that abstract away its complexity often abstract away its essence.Mathematics and the physical sciences made great strides for three centuries by constructing simplified models of complex phenomena, deriving properties from the models, and verifying those properties experimentally. This worked because the complexities（75）in the models were not the essential properties of the phenomena. It does not work when the complexities are the essence.Many of the classical problems of developing software products derive from this essential complexity and its nonlinear increases with size. Not only technical problems but management problems as well come from the complexity.",
		"a1": "A. linear",
		"a2": "B. nonlinear",
		"a3": "C. parallel",
		"a4": "D. additive",
		"sl": "B"
	}, {
		"tihao": 74,
		"q": "Software entities are more complex for their size than perhaps any other human construct, because no two parts are alike (at least above the statement level). If they are, we make the two similar parts into one, a（71）, open or closed. In this respect software systems differ profoundly from computers,buildings, or automobiles, where repeated elements abound.Digital computers are themselves more complex than most things people build; they have very large numbers of states. This makes conceiving, describing, and testing them hard. Software systems have orders of magnitude more （72）than computers do.Likewise, a scaling-up of a software entity is not merely a repetition of the same elements in larger size; it is necessarily an increase in the number of different elements. In most cases, the elements interact with each other in some（73）fashion,and the complexity of the whole increases much more than linearly.The complexity of software is a(an)（74）property, not an accidental one. Hence descriptions of a software entity that abstract away its complexity often abstract away its essence.Mathematics and the physical sciences made great strides for three centuries by constructing simplified models of complex phenomena, deriving properties from the models, and verifying those properties experimentally. This worked because the complexities（75）in the models were not the essential properties of the phenomena. It does not work when the complexities are the essence.Many of the classical problems of developing software products derive from this essential complexity and its nonlinear increases with size. Not only technical problems but management problems as well come from the complexity.",
		"a1": "A. surface",
		"a2": "B. outside",
		"a3": "C. exterior",
		"a4": "D. essential",
		"sl": "D"
	}, {
		"tihao": 75,
		"q": "Software entities are more complex for their size than perhaps any other human construct, because no two parts are alike (at least above the statement level). If they are, we make the two similar parts into one, a（71）, open or closed. In this respect software systems differ profoundly from computers,buildings, or automobiles, where repeated elements abound.Digital computers are themselves more complex than most things people build; they have very large numbers of states. This makes conceiving, describing, and testing them hard. Software systems have orders of magnitude more （72）than computers do.Likewise, a scaling-up of a software entity is not merely a repetition of the same elements in larger size; it is necessarily an increase in the number of different elements. In most cases, the elements interact with each other in some（73）fashion,and the complexity of the whole increases much more than linearly.The complexity of software is a(an)（74）property, not an accidental one. Hence descriptions of a software entity that abstract away its complexity often abstract away its essence.Mathematics and the physical sciences made great strides for three centuries by constructing simplified models of complex phenomena, deriving properties from the models, and verifying those properties experimentally. This worked because the complexities（75）in the models were not the essential properties of the phenomena. It does not work when the complexities are the essence.Many of the classical problems of developing software products derive from this essential complexity and its nonlinear increases with size. Not only technical problems but management problems as well come from the complexity.",
		"a1": "A. fixed",
		"a2": "B. included",
		"a3": "C. ignored",
		"a4": "D. stabilized",
		"sl": "C"
	}]
}

软件设计13年上={}
软件设计13年下={
	"code": 0,
	"list": [{
		"tihao": 1,
		"q": "在程序执行过程中，Cache与主存的地址映像由(1)。",
		"a1": "A. 硬件自动完成",
		"a2": "B. 程序员调度",
		"a3": "C. 操作系统管理",
		"a4": "D. 程序员与操作系统协同完成",
		"sl": "A"
	}, {
		"tihao": 2,
		"q": "指令寄存器的位数取决于(2)。",
		"a1": "A. 存储器的容量",
		"a2": "B. 指令字长",
		"a3": "C. 数据总线的宽度",
		"a4": "D. 地址总线的宽度",
		"sl": "B"
	}, {
		"tihao": 3,
		"q": "若计算机存储数据采用的是双符号位(00表示正号、11表示负号)，两个符号相同的数相加时，如果运算结果的两个符号位经(3)运算得1，则可断定这两个数相加的结果产生了溢出。",
		"a1": "A. 逻辑与",
		"a2": "B. 逻辑或",
		"a3": "C. 逻辑同或",
		"a4": "D. 逻辑异或",
		"sl": "D"
	}, {
		"tihao": 4,
		"q": "某指令流水线由4段组成，各段所需要的时间如下图所示。连续输入8条指令时的吞吐率(单位时间内流水线所完成的任务数或输出的结果数)为(4)。",
		"a1": "A. 8/56△t",
		"a2": "B. 8/32△t",
		"a3": "C. 8/28△t",
		"a4": "D. 8/24△t",
		"sl": "C"
	}, {
		"tihao": 5,
		"q": "(5)不是RISC的特点。",
		"a1": "A. 指令种类丰富",
		"a2": "B. 高效的流水线操作",
		"a3": "C. 寻址方式较少",
		"a4": "D. 硬布线控制",
		"sl": "A"
	}, {
		"tihao": 6,
		"q": "若某计算机字长为32位，内存容量为2GB,按字编址，则可寻址范围为(6)。",
		"a1": "A. 1024M",
		"a2": "B. 1GB",
		"a3": "C. 512M",
		"a4": "D. 2GB",
		"sl": "C"
	}, {
		"tihao": 7,
		"q": "下列网络攻击行为中，属于DoS攻击的是(7)。",
		"a1": "A. 特洛伊木马攻击",
		"a2": "B. SYN Flooding攻击",
		"a3": "C. 端口欺骗攻击",
		"a4": "D. IP欺骗攻击",
		"sl": "B"
	}, {
		"tihao": 8,
		"q": "PKI体制中，保证数字证书不被篡改的方法是(8)。",
		"a1": "A. 用CA的私钥对数字证书签名",
		"a2": "B. 用CA的公钥对数字证书签名",
		"a3": "C. 用证书主人的私钥对数字证书签名",
		"a4": "D. 用证书主人的公钥对数字证书签名",
		"sl": "A"
	}, {
		"tihao": 9,
		"q": "下列算法中，不属于公开密钥加密算法的是(9)。",
		"a1": "A. ECC",
		"a2": "B. DSA",
		"a3": "C. RSA",
		"a4": "D. DES",
		"sl": "D"
	}, {
		"tihao": 10,
		"q": "矢量图是常用的图形图像表示形式， (10)是描述矢量图的基本组成单位。",
		"a1": "A. 像素",
		"a2": "B. 像素点",
		"a3": "C. 图元",
		"a4": "D. 二进制位",
		"sl": "C"
	}, {
		"tihao": 11,
		"q": "视频信息是连续的图像序列，(11)是构成视频信息的基本单元。",
		"a1": "A. 帧",
		"a2": "B. 场",
		"a3": "C. 幅",
		"a4": "D. 像素",
		"sl": "A"
	}, {
		"tihao": 12,
		"q": "以下多媒体素材编辑软件中，(12)主要用于动画编辑和处理。",
		"a1": "A. WPS",
		"a2": "B. Xara3D",
		"a3": "C. PhotoShop",
		"a4": "D. Cool Edit Pro",
		"sl": "B"
	}, {
		"tihao": 13,
		"q": "为说明某一问题，在学术论文中需要引用某些资料。以下叙述中，(13)是不正确的。",
		"a1": "A. 既可引用发表的作品，也可引用未发表的作品",
		"a2": "B. 只能限于介绍、评论作品",
		"a3": "C. 只要不构成自己作品的主要部分，可适当引用资料",
		"a4": "D. 不必征得原作者的同意，不需要向他支付报酬",
		"sl": "A"
	}, {
		"tihao": 14,
		"q": "以下作品中，不适用或不受著作权法保护的是 (14)。",
		"a1": "A. 某教师在课堂上的讲课",
		"a2": "B. 某作家的作品《红河谷》",
		"a3": "C. 最高人民法院组织编写的《行政诉讼案例选编》",
		"a4": "D. 国务院颁布的《计算机软件保护条例》",
		"sl": "D"
	}, {
		"tihao": 15,
		"q": "以下关于数据流图中基本加工的叙述，不正确的是(15)。",
		"a1": "A. 对每一个基本加工，必须有一个加工规格说明",
		"a2": "B. 加工规格说明必须描述把输入数据流变换为输出数据流的加工规则",
		"a3": "C. 加工规格说明必须描述实现加工的具体流程",
		"a4": "D. 决策表可以用来表示加工规格说明",
		"sl": "C"
	}, {
		"tihao": 16,
		"q": "在划分模块时，一个模块的作用范围应该在其控制范围之内。若发现其作用范围不在其控制范围内，则(16)不是适当的处理方法。",
		"a1": "A. 将判定所在模块合并到父模块中，使判定处于较高层次",
		"a2": "B. 将受判定影响的模块下移到控制范围内",
		"a3": "C. 将判定上移到层次较高的位置",
		"a4": "D. 将父模块下移，使该判定处于较高层次",
		"sl": "D"
	}, {
		"tihao": 17,
		"q": "下图是一个软件项目的活动图，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，则里程碑(17)在关键路径上。若在实际项目进展中，活动AD在活动AC开始3天后才开始，而完成活动DG过程中，由于有临时事件发生，实际需要15天才能完成，则完成该项目的最短时间比原计划多了(18)天。",
		"a1": "A. B",
		"a2": "B. C",
		"a3": "C. D",
		"a4": "D. I",
		"sl": "B"
	}, {
		"tihao": 18,
		"q": "下图是一个软件项目的活动图，其中顶点表示项目里程碑，连接顶点的边表示包含的活动，则里程碑(17)在关键路径上。若在实际项目进展中，活动AD在活动AC开始3天后才开始，而完成活动DG过程中，由于有临时事件发生，实际需要15天才能完成，则完成该项目的最短时间比原计划多了(18)天。",
		"a1": "A. 8",
		"a2": "B. 3",
		"a3": "C. 5",
		"a4": "D. 6",
		"sl": "B"
	}, {
		"tihao": 19,
		"q": "针对《关键职员在项目未完成时就跳槽》的风险，最不合适的风险管理策略是(19)。",
		"a1": "A. 对每一个关键性的技术人员，要培养后备人员",
		"a2": "B. 建立项目组，以使大家都了解有关开发活动的信息",
		"a3": "C. 临时招聘具有相关能力的新职员",
		"a4": "D. 对所有工作组织细致的评审",
		"sl": "C"
	}, {
		"tihao": 20,
		"q": "程序运行过程中常使用参数在函数(过程)间传递信息，引用调用传递的是实参的(20)。",
		"a1": "A. 地址",
		"a2": "B. 类型",
		"a3": "C. 名称",
		"a4": "D. 值",
		"sl": "A"
	}, {
		"tihao": 21,
		"q": "己知文法G: S→A0|B1，A→S1|1, B→S0|0,其中S是开始符号。从S出发可以推导出(21)。",
		"a1": "A. 所有由0构成的字符串",
		"a2": "B. 所有由1构成的字符串",
		"a3": "C. 某些0和1个数相等的字符串",
		"a4": "D. 所有0和1个数不同的字符串",
		"sl": "C"
	}, {
		"tihao": 22,
		"q": "算术表达式a+(b-C)*d的后缀式是(22) (-、+、*表示算术的减、加、乘运算，运算符的优先级和结合性遵循惯例)。",
		"a1": "A. b c - d * a +",
		"a2": "B. a b c - d * +",
		"a3": "C. a b + c - d *",
		"a4": "D. a b c d - * +",
		"sl": "B"
	}, {
		"tihao": 23,
		"q": "假设系统采用PV操作实现进程同步与互斥，若有n个进程共享一台扫描仪，那么当信号量S的值为-3时，表示系统中有(23)个进程等待使用扫描仪。",
		"a1": "A. 0",
		"a2": "B. n-3",
		"a3": "C. 3",
		"a4": "D. n",
		"sl": "C"
	}, {
		"tihao": 24,
		"q": "假设段页式存储管理系统中的地址结构如下图所示，则系统中(24)。",
		"a1": "A. 页的大小为4K，每个段的大小均为4096个页，最多可有256个段",
		"a2": "B. 页的大小为4K,每个段最大允许有4096个页，最多可有256个段",
		"a3": "C. 页的大小为8K,每个段的大小均为2048个页，最多可有128个段",
		"a4": "D. 页的大小为8K,每个段最大允许有2048个页，最多可有128个段",
		"sl": "B"
	}, {
		"tihao": 25,
		"q": "某文件管理系统采用位示图(bitmap)记录磁盘的使用情况。如果系统的字长为32位，磁盘物理块的大小为4MB，物理块依次编号为：0、1、2、位示图字依次编号为：0、1、2、那么16385号物理块的使用情况在位示图中的第(25)个字中描述;如果磁盘的容量为1000GB,那么位示图需要(26)个字来表示。",
		"a1": "A. 128",
		"a2": "B. 256",
		"a3": "C. 512",
		"a4": "D. 1024",
		"sl": "C"
	}, {
		"tihao": 26,
		"q": "某文件管理系统采用位示图(bitmap)记录磁盘的使用情况。如果系统的字长为32位，磁盘物理块的大小为4MB，物理块依次编号为：0、1、2、位示图字依次编号为：0、1、2、那么16385号物理块的使用情况在位示图中的第(25)个字中描述;如果磁盘的容量为1000GB,那么位示图需要(26)个字来表示。",
		"a1": "A. 1200",
		"a2": "B. 3200",
		"a3": "C. 6400",
		"a4": "D. 8000",
		"sl": "D"
	}, {
		"tihao": 27,
		"q": "假设系统中有三类互斥资源R1、R2和R3,可用资源数分别为10、5和3。在T0时刻系统中有P1、P2、P3、P4和P5五个进程，这些进程对资源的最大需求量和已分配资源数如下表所示，此时系统剩余的可用资源数分别为(27)。如果进程按(28)序列执行，那么系统状态是安全的。",
		"a1": "A. 1、1和0",
		"a2": "B. 1、1和1",
		"a3": "C. 2、1和0",
		"a4": "D. 2、0和1",
		"sl": "D"
	}, {
		"tihao": 28,
		"q": "假设系统中有三类互斥资源R1、R2和R3,可用资源数分别为10、5和3。在T0时刻系统中有P1、P2、P3、P4和P5五个进程，这些进程对资源的最大需求量和已分配资源数如下表所示，此时系统剩余的可用资源数分别为(27)。如果进程按(28)序列执行，那么系统状态是安全的。",
		"a1": "A. P1→P2→P4→P5→P3",
		"a2": "B. P5→P2→P4→P3→P1",
		"a3": "C. P4→P2→P1→P5→P3",
		"a4": "D. P5→P1→P4→P2→P3",
		"sl": "B"
	}, {
		"tihao": 29,
		"q": "(29)开发过程模型最不适用于开发初期对软件需求缺乏准确全面认识的情况。",
		"a1": "A. 瀑布",
		"a2": "B. 演化",
		"a3": "C. 螺旋",
		"a4": "D. 增量",
		"sl": "A"
	}, {
		"tihao": 30,
		"q": "(30)不是增量式开发的优势。",
		"a1": "A. 软件可以快速地交付",
		"a2": "B. 早期的增量作为原型，从而可以加强对系统后续开发需求的理解",
		"a3": "C. 具有最高优先级的功能首先交付，随着后续的增量不断加入，这就使得更重要的功能得到更多的测试",
		"a4": "D. 很容易将客户需求划分为多个增量",
		"sl": "D"
	}, {
		"tihao": 31,
		"q": "在对程序质量进行评审时，模块结构是一个重要的评审项，评审内容中不包括(31)。",
		"a1": "A. 数据结构",
		"a2": "B. 数据流结构",
		"a3": "C. 控制流结构",
		"a4": "D. 模块结构与功能结构之间的对应关系",
		"sl": "A"
	}, {
		"tihao": 32,
		"q": "SEI能力成熟度模型(SEICMM)把软件开发企业分为5个成熟度级别，其中(32)重点关注产品和过程质量。",
		"a1": "A. 级别2:重复级",
		"a2": "B. 级别3:确定级",
		"a3": "C. 级别4:管理级",
		"a4": "D. 级别5:优化级",
		"sl": "C"
	}, {
		"tihao": 33,
		"q": "系统可维护性的评价指标不包括(33)。",
		"a1": "A. 可理解性",
		"a2": "B. 可测试性",
		"a3": "C. 可移植性",
		"a4": "D. 可修改性",
		"sl": "C"
	}, {
		"tihao": 34,
		"q": "逆向工程从源代码或U标代码中提取设计信息，通常在原软件生命周期的(34)阶段进行。",
		"a1": "A. 需求分析",
		"a2": "B. 软件设计",
		"a3": "C. 软件实现",
		"a4": "D. 软件维护",
		"sl": "D"
	}, {
		"tihao": 35,
		"q": "一个程序根据输入的年份和月份计算该年中该月的天数，输入参数包括年份(正整数)、月份(用1〜12表示)。若用等价类划分测试方法进行测试，则(35)不是一个合适的测试用例(分号后表示测试的输出)。",
		"a1": "A. (2013,1;31)",
		"a2": "B. (0,1;‘错误’)",
		"a3": "C. (0,13;‘错误’)",
		"a4": "D. (2000,-1;‘错误’)",
		"sl": "C"
	}, {
		"tihao": 36,
		"q": "(36)不是单元测试主要检查的内容。",
		"a1": "A. 模块接口",
		"a2": "B. 局部数据结构",
		"a3": "C. 全局数据结构",
		"a4": "D. 重要的执行路径",
		"sl": "C"
	}, {
		"tihao": 37,
		"q": "在领域类模型中不包含(37)。",
		"a1": "A. 属性",
		"a2": "B. 操作",
		"a3": "C. 关联",
		"a4": "D. 领域对象",
		"sl": "D"
	}, {
		"tihao": 38,
		"q": "在执行如下所示的UML活动图时，能同时运行的最大线程数为(38)。",
		"a1": "A. 4",
		"a2": "B. 3",
		"a3": "C. 2",
		"a4": "D. 1",
		"sl": "C"
	}, {
		"tihao": 39,
		"q": "下图所示的UML序列图中，(39)表示返回消息，Account应该实现的方法有(40)。",
		"a1": "A. xfer",
		"a2": "B. check",
		"a3": "C. evaluation",
		"a4": "D. minus",
		"sl": "C"
	}, {
		"tihao": 40,
		"q": "下图所示的UML序列图中，(39)表示返回消息，Account应该实现的方法有(40)。",
		"a1": "A. xfer()",
		"a2": "B. xfer()、plus()和minus()",
		"a3": "C. check()、plus()和minus()",
		"a4": "D. xfer()、evaluation()、plus()和minus()",
		"sl": "B"
	}, {
		"tihao": 41,
		"q": "在面向对象技术中，(41)定义了超类和子类之间的关系，子类中以更具体的方式实现从父类继承来的方法称为(42)，不同类的对象通过(43)相互通信。",
		"a1": "A. 覆盖",
		"a2": "B. 继承",
		"a3": "C. 信息",
		"a4": "D. 多态",
		"sl": "B"
	}, {
		"tihao": 42,
		"q": "在面向对象技术中，(41)定义了超类和子类之间的关系，子类中以更具体的方式实现从父类继承来的方法称为(42)，不同类的对象通过(43)相互通信。",
		"a1": "A. 覆盖",
		"a2": "B. 继承",
		"a3": "C. 消息",
		"a4": "D. 多态",
		"sl": "A"
	}, {
		"tihao": 43,
		"q": "在面向对象技术中，(41)定义了超类和子类之间的关系，子类中以更具体的方式实现从父类继承来的方法称为(42)，不同类的对象通过(43)相互通信。",
		"a1": "A. 覆盖",
		"a2": "B. 继承",
		"a3": "C. 消息",
		"a4": "D. 多态",
		"sl": "C"
	}, {
		"tihao": 44,
		"q": "(44)设计模式定义一系列算法，把它们一个个封装起来，并且使它们可相互替换。这一模式使得算法可独立于它的客户而变化。",
		"a1": "A. 策略(Strategy)",
		"a2": "B. 抽象工厂(Abstract Factory)",
		"a3": "C. 观察者(Visitor)",
		"a4": "D. 状态(State)",
		"sl": "A"
	}, {
		"tihao": 45,
		"q": "在发布-订阅(Publish-Subscribe)消息模型中，订阅者订阅一个主题后，当该主题有新消息到达时，所有订阅者都会收到通知。(45)设计模式最适合这一模型。",
		"a1": "A. 适配器(Adapter)",
		"a2": "B. 通知(Notifier)",
		"a3": "C. 状态(State)",
		"a4": "D. 观察者(Observer)",
		"sl": "D"
	}, {
		"tihao": 46,
		"q": "下图所示为(46)设计模式，适用于：(47)。",
		"a1": "A. 组件(Component)",
		"a2": "B. 适配器(Adapter)",
		"a3": "C. 组合(Composite)",
		"a4": "D. 装饰器(Decorator)",
		"sl": "C"
	}, {
		"tihao": 47,
		"q": "下图所示为(46)设计模式，适用于：(47)。",
		"a1": "A. 表示对象的部分-整体层次结构",
		"a2": "B. 不希望在抽象和它的实现部分之间有一个固定的绑定关系",
		"a3": "C. 在不影响其他对象的情况下，以动态、透明的方式给单个对象添加职责",
		"a4": "D. 使所有接口不兼容类可以一起工作",
		"sl": "A"
	}, {
		"tihao": 48,
		"q": "将高级语言程序翻译为机器语言程序的过程中，常引入中间代码，其好处是(48)。",
		"a1": "A. 有利于进行反编译处理",
		"a2": "B. 有利于进行与机器无关的优化处理",
		"a3": "C. 尽早发现语法错误",
		"a4": "D. 可以简化语法和语义分析",
		"sl": "B"
	}, {
		"tihao": 49,
		"q": "对高级语言源程序进行编译的过程中，有限自动机(NFA或DFA)是进行(49)的适当工具。",
		"a1": "A. 词法分析",
		"a2": "B. 语法分析",
		"a3": "C. 语义分析",
		"a4": "D. 出错处理",
		"sl": "A"
	}, {
		"tihao": 50,
		"q": "弱类型语言(动态类型语言)是指不需要进行变量/对象类型声明的语言。(50)属于弱类型语言。",
		"a1": "A. Java",
		"a2": "B. C/C++",
		"a3": "C. Python",
		"a4": "D. C#",
		"sl": "C"
	}, {
		"tihao": 51,
		"q": "若有关系R(A,B,C,D,E)和S(B,C,F,G),则R与S自然联结运算后的属性列有(51)个，与表达式π1,3,6,7(σ3<6(RS))等价的SQL语句如下：SELECT (52) FROM (53) WHERE (54);",
		"a1": "A. 5",
		"a2": "B. 6",
		"a3": "C. 7",
		"a4": "D. 9",
		"sl": "C"
	}, {
		"tihao": 52,
		"q": "若有关系R(A,B,C,D,E)和S(B,C,F,G),则R与S自然联结运算后的属性列有(51)个，与表达式π1,3,6,7(σ3<6(RS))等价的SQL语句如下：SELECT (52) FROM (53) WHERE (54);",
		"a1": "A. A,R.C,F,G",
		"a2": "B. A,C,S.B,S.F",
		"a3": "C. A,C,S.B,S.C",
		"a4": "D. C.R.A, R.C, S.B,S.C",
		"sl": "A"
	}, {
		"tihao": 53,
		"q": "若有关系R(A,B,C,D,E)和S(B,C,F,G),则R与S自然联结运算后的属性列有(51)个，与表达式π1,3,6,7(σ3<6(RS))等价的SQL语句如下：SELECT (52) FROM (53) WHERE (54);",
		"a1": "A. R",
		"a2": "B. S",
		"a3": "C. RS",
		"a4": "D. R, S",
		"sl": "D"
	}, {
		"tihao": 54,
		"q": "若有关系R(A,B,C,D,E)和S(B,C,F,G),则R与S自然联结运算后的属性列有(51)个，与表达式π1,3,6,7(σ3<6(RS))等价的SQL语句如下：SELECT (52) FROM (53) WHERE (54);",
		"a1": "A. R.B=S.BANDR.C=S.CANDR.C<S.B",
		"a2": "B. R.B=S.BANDR.C=S.CANDR.C<S.F",
		"a3": "C. C.R.B=S.BORR.C=S.CORR.C<S.B",
		"a4": "D. R.B=S.BORR.C=S.CORR.C<S.F",
		"sl": "B"
	}, {
		"tihao": 55,
		"q": "在分布式数据库系统中，(55)是指用户无需知道数据存放的物理位置。",
		"a1": "A. 分片透明",
		"a2": "B. 复制透明",
		"a3": "C. 逻辑透明",
		"a4": "D. 位置透明",
		"sl": "D"
	}, {
		"tihao": 56,
		"q": "计算机系统的软硬件故障可能会造成数据库中的数据被破坏。为了防止这一问题，通常需要(56),以便发生故障时恢复数据库。",
		"a1": "A. 定期安装DBMS和应用程序",
		"a2": "B. 定期安装应用程序，并将数据库做镜像",
		"a3": "C. 定期安装DBMS，并将数据库作备份",
		"a4": "D. 定期将数据库作备份；在进行事务处理时，需要将数据更新写入日志文件",
		"sl": "D"
	}, {
		"tihao": 57,
		"q": "以下关于线性表存储结构的叙述，正确的是(57)。",
		"a1": "A. 线性表采用顺序存储结构时，访问表中任意一个指定序号元素的时间复杂度为常量级",
		"a2": "B. 线性表采用顺序存储结构时，在表中任意位置插入新元素的运算时间复杂度为常量级",
		"a3": "C. 线性表采用链式存储结构时，访问表中任意一个指定序号元素的时间复杂度为常量级",
		"a4": "D. 线性表采用链式存储结构时，在表中任意位置插入新元素的运算时间复杂度为常量级",
		"sl": "A"
	}, {
		"tihao": 58,
		"q": "设循环队列Q的定义中有front和size两个域变量，其中front表示队头元素的指针，size表示队列的长度，如下图所示(队列长度为3,队头元素为x、队尾元素为z)。设队列的存储空间容量为M，则队尾元素的指针为(58)。",
		"a1": "A. (Q.front+Q.size-1)",
		"a2": "B. (Q.front+Q.size-1+M)%M",
		"a3": "C. (Q.front-Q.size)",
		"a4": "D. (Q.front-Q.size+M)%M",
		"sl": "B"
	}, {
		"tihao": 59,
		"q": "在一个有向图G的拓扑序列中，顶点Vi排列在Vj之前，说明图G中(59)。",
		"a1": "A. 一定存在弧＜vi,vj＞",
		"a2": "B. 一定存在弧＜vj,vi＞",
		"a3": "C. 可能存在vi到vj的路径，而不可能存在Vj到vi的路径",
		"a4": "D. 可能存在vj到vi的路径，而不可能存在vi到vj的路径",
		"sl": "C"
	}, {
		"tihao": 60,
		"q": "以下关于哈夫曼树的叙述，正确的是(60)。",
		"a1": "A. 哈夫曼树一定是满二叉树，其每层结点数都达到最大值",
		"a2": "B. 哈夫曼树一定是平衡二叉树，其每个结点左右子树的高度差为-1、0或1",
		"a3": "C. 哈夫曼树中左孩子结点的权值小于父结点、右孩子结点的权值大于父结点",
		"a4": "D. 哈夫曼树中叶子结点的权值越小则距离树根越远、叶子结点的权值越大则距离树根越近",
		"sl": "D"
	}, {
		"tihao": 61,
		"q": "某哈希表(散列表)的长度为n，设散列函数为H(Key)=Key mod p，采用线性探测法解决冲突。以下关于p值的叙述中，正确的是(61)。",
		"a1": "A. p的值一般为不大于n且最接近n的质数",
		"a2": "B. p的值一般为大于n的任意整数",
		"a3": "C. p的值必须为小于n的合数",
		"a4": "D. p的值必须等于n",
		"sl": "A"
	}, {
		"tihao": 62,
		"q": "对n个基本有序的整数进行排序，若采用插入排序算法，则时间和空间复杂度分别为(62);若采用快速排序算法，则时间和空间复杂度分别为(63)。",
		"a1": "A. O(n2)和O(n)",
		"a2": "B. O(n)和O(n)",
		"a3": "C. O(n2)和O(1)",
		"a4": "D. O(n)和O(1)",
		"sl": "D"
	}, {
		"tihao": 63,
		"q": "对n个基本有序的整数进行排序，若采用插入排序算法，则时间和空间复杂度分别为(62);若采用快速排序算法，则时间和空间复杂度分别为(63)。",
		"a1": "A. O(n2)和O(n)",
		"a2": "B. O(nlgn)和O(n)",
		"a3": "C. O(n2)和O(1)",
		"a4": "D. O(nlgn)和O(1)",
		"sl": "C"
	}, {
		"tihao": 64,
		"q": "在求解某问题时，经过分析发现该问题具有最优子结构性质，求解过程中子问题被重复求解，则采用(64)算法设计策略;若定义问题的解空间，以深度优先的方式搜索解空间，则采用(65)算法设计策略。",
		"a1": "A. 分治",
		"a2": "B. 动态规划",
		"a3": "C. 贪心",
		"a4": "D. 回溯",
		"sl": "B"
	}, {
		"tihao": 65,
		"q": "在求解某问题时，经过分析发现该问题具有最优子结构性质，求解过程中子问题被重复求解，则采用(64)算法设计策略;若定义问题的解空间，以深度优先的方式搜索解空间，则采用(65)算法设计策略。",
		"a1": "A. 动态规划",
		"a2": "B. 贪心",
		"a3": "C. 回溯",
		"a4": "D. 分支限界",
		"sl": "C"
	}, {
		"tihao": 66,
		"q": "某单位的局域网配置如下图所示，PC2发送到Internet上的报文的源IP地址为(66)。",
		"a1": "A. 192.168.0.2",
		"a2": "B. 192.168.0.1",
		"a3": "C. 202.117.112.1",
		"a4": "D. 202.117.112.2",
		"sl": "D"
	}, {
		"tihao": 67,
		"q": "在IPv4向IPv6过渡期间，如果要使得两个IPv6结点可以通过现有的IPv4网络进行通信，则应该使用(67);如果要使得纯IPv6结点可以与纯IPv4结点进行通信，则需要使用(68)。",
		"a1": "A. 堆栈技术",
		"a2": "B. 双协议栈技术",
		"a3": "C. 隧道技术",
		"a4": "D. 翻译技术",
		"sl": "C"
	}, {
		"tihao": 68,
		"q": "在IPv4向IPv6过渡期间，如果要使得两个IPv6结点可以通过现有的IPv4网络进行通信，则应该使用(67);如果要使得纯IPv6结点可以与纯IPv4结点进行通信，则需要使用(68)。",
		"a1": "A. 堆栈技术",
		"a2": "B. 双协议桟技术",
		"a3": "C. 隧道技术",
		"a4": "D. 翻译技术",
		"sl": "D"
	}, {
		"tihao": 69,
		"q": "POP3协议采用(69)模式进行通信,当客户机需要服务时，客户端软件与POP3服务器建立(70)连接。",
		"a1": "A. Browser/Server",
		"a2": "B. Client/Server",
		"a3": "C. PeertoPeer",
		"a4": "D. PeertoServer",
		"sl": "B"
	}, {
		"tihao": 70,
		"q": "POP3协议采用(69)模式进行通信,当客户机需要服务时，客户端软件与POP3服务器建立(70)连接。",
		"a1": "A. TCP",
		"a2": "B. UDP",
		"a3": "C. PHP",
		"a4": "D. IP",
		"sl": "A"
	}, {
		"tihao": 71,
		"q": "There is nothing in this world constant but inconstancy. —SWIFT Project after project designs a set of algorithms and then plunges into construction of customer-deliverable software on a schedule that demands delivery of the first thing built.In most projects, the first system built is (71) usable. It may be too slow，too big, awkward to use, or all three. There is no (72) but to start again, smarting but smarter, and build a redesigned version in which these problems are solved. The discard and (73) may be done in one lump, or it may be done piece-by-piece. But all large-system experience shows that it will be done. Where a new system concept or new technology is used, one has to build a system to throw away, for even the best planning is not so omniscient (全知的）as to get it right the first time.The management question, therefore, is not whether to build a pilot system and throw it away. You will do that. The only question is whether to plan in advance to build a (74) , or to promise to deliver the throwaway to customers. Seen this way, the answer is mi.ch clearer. Delivering that throwaway to customers buys time, but it does so only at the (75) of agony (极大痛苦)for the user, distraction for the builders while they do the redesign, and a bad reputation for the product that the best redesign will find hard to live down.",
		"a1": "A. almost",
		"a2": "B. often",
		"a3": "C. usually",
		"a4": "D. barely",
		"sl": "D"
	}, {
		"tihao": 72,
		"q": "There is nothing in this world constant but inconstancy. —SWIFT Project after project designs a set of algorithms and then plunges into construction of customer-deliverable software on a schedule that demands delivery of the first thing built.In most projects, the first system built is (71) usable. It may be too slow，too big, awkward to use, or all three. There is no (72) but to start again, smarting but smarter, and build a redesigned version in which these problems are solved. The discard and (73) may be done in one lump, or it may be done piece-by-piece. But all large-system experience shows that it will be done. Where a new system concept or new technology is used, one has to build a system to throw away, for even the best planning is not so omniscient (全知的）as to get it right the first time.The management question, therefore, is not whether to build a pilot system and throw it away. You will do that. The only question is whether to plan in advance to build a (74) , or to promise to deliver the throwaway to customers. Seen this way, the answer is mi.ch clearer. Delivering that throwaway to customers buys time, but it does so only at the (75) of agony (极大痛苦)for the user, distraction for the builders while they do the redesign, and a bad reputation for the product that the best redesign will find hard to live down.",
		"a1": "A. alternative",
		"a2": "B. need",
		"a3": "C. possibility",
		"a4": "D. solution",
		"sl": "A"
	}, {
		"tihao": 73,
		"q": "There is nothing in this world constant but inconstancy. —SWIFT Project after project designs a set of algorithms and then plunges into construction of customer-deliverable software on a schedule that demands delivery of the first thing built.In most projects, the first system built is (71) usable. It may be too slow，too big, awkward to use, or all three. There is no (72) but to start again, smarting but smarter, and build a redesigned version in which these problems are solved. The discard and (73) may be done in one lump, or it may be done piece-by-piece. But all large-system experience shows that it will be done. Where a new system concept or new technology is used, one has to build a system to throw away, for even the best planning is not so omniscient (全知的）as to get it right the first time.The management question, therefore, is not whether to build a pilot system and throw it away. You will do that. The only question is whether to plan in advance to build a (74) , or to promise to deliver the throwaway to customers. Seen this way, the answer is mi.ch clearer. Delivering that throwaway to customers buys time, but it does so only at the (75) of agony (极大痛苦)for the user, distraction for the builders while they do the redesign, and a bad reputation for the product that the best redesign will find hard to live down.",
		"a1": "A. design",
		"a2": "B. redesign",
		"a3": "C. plan",
		"a4": "D. build",
		"sl": "B"
	}, {
		"tihao": 74,
		"q": "There is nothing in this world constant but inconstancy. —SWIFT Project after project designs a set of algorithms and then plunges into construction of customer-deliverable software on a schedule that demands delivery of the first thing built.In most projects, the first system built is (71) usable. It may be too slow，too big, awkward to use, or all three. There is no (72) but to start again, smarting but smarter, and build a redesigned version in which these problems are solved. The discard and (73) may be done in one lump, or it may be done piece-by-piece. But all large-system experience shows that it will be done. Where a new system concept or new technology is used, one has to build a system to throw away, for even the best planning is not so omniscient (全知的）as to get it right the first time.The management question, therefore, is not whether to build a pilot system and throw it away. You will do that. The only question is whether to plan in advance to build a (74) , or to promise to deliver the throwaway to customers. Seen this way, the answer is mi.ch clearer. Delivering that throwaway to customers buys time, but it does so only at the (75) of agony (极大痛苦)for the user, distraction for the builders while they do the redesign, and a bad reputation for the product that the best redesign will find hard to live down.",
		"a1": "A. throwaway",
		"a2": "B. system",
		"a3": "C. software",
		"a4": "D. product",
		"sl": "A"
	}, {
		"tihao": 75,
		"q": "There is nothing in this world constant but inconstancy. —SWIFT Project after project designs a set of algorithms and then plunges into construction of customer-deliverable software on a schedule that demands delivery of the first thing built.In most projects, the first system built is (71) usable. It may be too slow，too big, awkward to use, or all three. There is no (72) but to start again, smarting but smarter, and build a redesigned version in which these problems are solved. The discard and (73) may be done in one lump, or it may be done piece-by-piece. But all large-system experience shows that it will be done. Where a new system concept or new technology is used, one has to build a system to throw away, for even the best planning is not so omniscient (全知的）as to get it right the first time.The management question, therefore, is not whether to build a pilot system and throw it away. You will do that. The only question is whether to plan in advance to build a (74) , or to promise to deliver the throwaway to customers. Seen this way, the answer is mi.ch clearer. Delivering that throwaway to customers buys time, but it does so only at the (75) of agony (极大痛苦)for the user, distraction for the builders while they do the redesign, and a bad reputation for the product that the best redesign will find hard to live down.",
		"a1": "A. worth",
		"a2": "B. value",
		"a3": "C. cost",
		"a4": "D. invaluable",
		"sl": "C"
	}]
}



# 遍历所有节点
def walkData(rootNode, levle, resultList):
    global uniqueId
    tempList = [levle, rootNode.tag]
    resultList.append(tempList)
    uniqueId += 1

    # 遍历所有子节点
    childrenNode = rootNode.getchildren();
    if len(childrenNode) == 0:
        return
    for child in childrenNode:
        walkData(child, levle + 1, resultList)
    return


def getXmlData(path):
    # 节点的深度从 1 开始
    level = 1
    reusltList = []

    # 循环遍历所有文件
    for root, dirs, files in os.walk(path):
        i = 0;
        while len(files) > i:
            a, d, dx = AnalyzeAPK(path + "\\" + files[i])
            rootNode = AXMLPrinter.get_xml_obj(a.get_android_manifest_axml());
            walkData(rootNode, level, reusltList)
            return reusltList
            i = i + 1;

if __name__ == '__main__':
    # url = 'https://www.rkstw.com/api/xuanze/485' 软件设计17年上
    # url='https://www.rkstw.com/api/xuanze/91'软件设计13年下
    url='https://www.rkstw.com/api/xuanze/420'

    web_data = requests.get(url)

    web_data.encoding = 'utf-8'

    soup = BeautifulSoup(web_data.text, 'lxml')
    # 将python字典类型变成json数据格式
    # tj=json.dumps(soup.text)
    # ttttt=json.loads(tj)
    # print(type(ttttt))
    t1 = 软件设计16年下['list']
# data = pd.read_excel('试题数据导入模板.xlsx', engine='openpyxl')
# print(data)  # 打印显示表格的属性，几行几列
wb = load_workbook("试题数据导入模板.xlsx")  # 生成一个已存在的wookbook对象
wb1 = wb.active  # 激活sheet
row=3
column=0
for kv in t1:
    # wb1.cell(2, 2, 'pass2')  # 往sheet中的第二行第二列写入‘pass2’的数据
    # 题干
    wb1.cell(row, 4, kv['q'])
    # 选项A
    wb1.cell(row, 6, kv['a1'])
    # 选项B
    wb1.cell(row, 7, kv['a2'])
    # 选项C
    wb1.cell(row, 8, kv['a3'])
    # 选项D
    wb1.cell(row, 9, kv['a4'])
    # 答案
    wb1.cell(row, 10, kv['sl'])
    row = row + 1
pass

# 保存
wb.save("试题数据导入模板.xlsx")  # 保存


# # 创建 excel 指令：pip install xlwt、pip install xlrd
# workbook = xlwt.Workbook(encoding='utf-8')
# # 新建工作簿 sheet
# sheet = workbook.add_sheet("androidPermission")
# dirPath='C:\\Users\\rgao\\Desktop\\workSpace\\python\\permiss\\app'
# # 得到所有节点
# allDatas=getXmlData(dirPath);
# # 去重
# notDuplicateData=[]
# for element in allDatas:
#     if(element not in notDuplicateData):
#         notDuplicateData.append(element)
#
# # 和所有节点作对比，得到同深度节点数
# for j,newData in enumerate (notDuplicateData):
#     # 节点个数
#     i=0
#     for data in allDatas:
#         # 同深度、同标签名的情况
#         if(newData[0] == data[0] and newData[1]==data[1]):
#             i=i+1
#     notDuplicateData[j].append(i)
# pass
#
# row = 0
# column = 0
#
# # 总是用当前深度和上一个深度作比较，得到百分比
# for newData in notDuplicateData:
#     # 百分比
#     o = 0
#     for j,data in enumerate(notDuplicateData):
#         if (newData[0]+1 ==  data[0]):
#             o = str(newData[1])+'/'+str(data[1])+'{:.2%}'.format(newData[2]/data[2])
#
#     # 标签名
#     sheet.write(row, column, newData[1])
#     # 个数
#     sheet.write(row + 1, column, newData[2])
#     # 层级数
#     sheet.write(row + 2, column, newData[0])
#     # 百分比
#     sheet.write(row+3,column,o)
#     column = column + 1
# pass
#
# # 保存
# workbook.save("allXml.xls")
